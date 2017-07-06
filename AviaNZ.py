# Interface.py
#
# This is the main class for the AviaNZ interface
# It's fairly simple, but seems to work OK
# Version 0.10 16/04/17
# Author: Stephen Marsland, with input from Nirosha Priyadarshani

#    AviaNZ birdsong analysis program
#    Copyright (C) 2017

#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.

import sys, os, json, platform    #,glob
from PyQt4.QtCore import *
from PyQt4.QtGui import *
#from PyQt4.QtWebKit import *
import PyQt4.phonon as phonon

import wavio
from scipy.misc import imread, imsave
#import librosa as lr
import numpy as np

import pyqtgraph as pg
pg.setConfigOption('background','w')
pg.setConfigOption('foreground','k')
pg.setConfigOption('antialias',True)
from pyqtgraph.Qt import QtCore, QtGui
from pyqtgraph.dockarea import *
import pyqtgraph.functions as fn

import SupportClasses as SupportClasses
import Dialogs as Dialogs
import SignalProc
import Segment
#import Features
#import Learning
import interface_FindSpecies
import WaveletSegment
#import math

from openpyxl import load_workbook, Workbook

#import pyqtgraph.parametertree.parameterTypes as pTree
# ==============
# TODO

# (4) pysox

# Check that the paging works, add in a label to say where you are up to (where to put it?!)

# Good idea to project disconnects with exceptions

# Finish segmentation
#   Mostly there, need to test them
#   Add a minimum length of time for a segment -> make this a parameter
#   Finish sorting out parameters for median clipping segmentation, energy segmentation
#   Finish cross-correlation to pick out similar bits of spectrogram -> and what other methods?
#   Add something that aggregates them -> needs planning

# Integrate the wavelet segmentation
    # Remove the code that is in SignalProc and use that one

# At times the program does not respond and ask to repair/close (e.g. when move the overview slider fast or something like that).
# Need to work on memory management!

# Interface -> inverted spectrogram does not work - spec and amp do not synchronize

# Actions -> Denoise -> median filter check
# Make the median filter on the spectrogram have params and a dialog. Other options?

# Fundamental frequency
#   Smoothing?
#   Add shape metric
#   Try the harvest f0
#   Try yaapt or bana (tried yaapt, awful)

# Finish the raven features

# Add in the wavelet segmentation for kiwi, ruru
# Think about nice ways to train them

# Would it be good to smooth the image? Actually, lots of ideas here! Might be nice way to denoise?
    # Median filter, smoothing, consider also grab-cut
    # Continue to play with inverting spectrogram

# Colourmaps
    # HistogramLUTItem

# Make the scrollbar be the same size as the spectrogram -> hard!


# Context menu different for day and night birds?

# The ruru file is a good one to play with for now

# Look into ParameterTree for saving the config stuff in particular

# Needs decent testing

# Minor:
# Consider always resampling to 22050Hz (except when it's less in file :) )?
# Font size to match segment size -> make it smaller, could also move it up or down as appropriate
# Where should label be written?
# Use intensity of colour to encode certainty?
# If don't select something in context menu get error -> not critical
# Colours of the segments to be visible with different colourmaps? Not important!

# Look at raven and praat and luscinia -> what else is actually useful? Other annotations on graphs?

# Given files > 5 mins, split them into 5 mins versions anyway (code is there, make it part of workflow)
# Don't really want to load the whole thing, just 5 mins, and then move through with arrows -> how?
# This is sometimes called paging, I think. (y, sr = librosa.load(filename, offset=15.0, duration=5.0) might help. Doesn't do much for the overview through)
# I tried saving a figure of the overview, but they are too big in general

# Things to remember
    # When adding new classes, make sure to pass new data to them in undoing and loading

# This version has the selection of birds using a context menu and then has removed the radio buttons
# Code is still there, though, just commented out. Add as an option?

# Diane:
    # menu
    # for matching, show matched segment, and extended 'guess' (with some overlap)
    # Something to show nesting of segments, such as a number of segments in the top bit
    # Find similar segments in other files -- other birds
    # Group files by species

# Rebecca:
    # x colour spectrogram
    # x add a marker on the overview to show where you have marked segments, with different colours for unknown, possible
    # x reorder the list dynamically by amount of use -> done, but maybe it should be an option?
    # Maybe include day or night differently in the context menu
    # x have a hot key to add the same bird repeatedly
    # Change the visible window width (or just add) magnify/shrink buttons
    # x Add date, time, person, location, would be good to have weather info calculated automatically (wind, rain -> don't bother), broken sound recorder

    # pull out any bird call (useful for if you don't know what a bird sounds like (Fiji petrel) or wind farm monitoring)
    # do bats!
    # Look up David Bryden (kokako data) looking at male, female, juvenile
    # Get all calls of a species
    # Look up freebird and raven and also BatSearch
# ===============

class AviaNZ(QMainWindow):
    """Main class for the user interface.
    Contains most of the user interface and plotting code"""

    def __init__(self,root=None,configfile=None,DOC=True,username=None):
        """Initialisation of the class. Load a configuration file, or create a new one if it doesn't
        exist. Also initialises the data structures and loads an initial file (specified explicitly)
        and sets up the window.
        One interesting configuration point is the DOC setting, which hides the more 'research' functions."""
        super(AviaNZ, self).__init__()
        self.root = root
        if configfile is not None:
            try:
                self.config = json.load(open(configfile))
                self.saveConfig = False
            except:
                print("Failed to load config file")
                self.genConfigFile()
                self.saveConfig = True
            self.configfile = configfile
        else:
            self.genConfigFile()
            self.saveConfig=True
            self.configfile = 'AviaNZconfig.txt'

        self.username = username

        # The data structures for the segments
        self.listLabels = []
        self.listRectanglesa1 = []
        self.listRectanglesa2 = []
        self.SegmentRects = []
        self.segmentPlots=[]
        self.box1id = -1
        self.DOC=DOC
        self.started=False
        self.bar = pg.InfiniteLine(angle=90, movable=True, pen={'color': 'c', 'width': 3})

        self.lastSpecies = "Don't Know"
        self.nFileSections = None
        self.resetStorageArrays()

        self.dirName = self.config['dirpath']
        self.previousFile = None
        self.focusRegion = None

        # Make the window and associated widgets
        QMainWindow.__init__(self, root)
        self.setWindowTitle('AviaNZ')

        # Make life easier for now: preload a birdsong
        firstFile = 'tril1.wav' #'male1.wav' # 'kiwi.wav'#'
        #self.firstFile = 'kiwi.wav'

        self.createMenu()
        self.createFrame()

        # Some safety checking for paths and files
        if not os.path.isdir(self.dirName):
            print("Directory doesn't exist: making it")
            os.makedirs(self.dirName)
        if not os.path.isfile(self.dirName+'/'+firstFile):
            fileName = QtGui.QFileDialog.getOpenFileName(self, 'Choose File', self.dirName, "Wav files (*.wav)")
            if fileName:
                firstFile = fileName
        self.fillFileList(firstFile)
        self.listLoadFile(QString(firstFile))
        #self.previousFile = firstFile

        # Save the segments every minute
        self.timer = QTimer()
        QObject.connect(self.timer, SIGNAL("timeout()"), self.saveSegments)
        self.timer.start(self.config['secsSave']*1000)

    def createMenu(self):
        """ Create the menu entries at the top of the screen and link them as appropriate.
        Some of them are initialised according to the data in the configuration file."""

        fileMenu = self.menuBar().addMenu("&File")
        fileMenu.addAction("&Open sound file", self.openFile, "Ctrl+O")
        # fileMenu.addAction("&Change Directory", self.chDir)
        fileMenu.addAction("&Set Operator/Reviewer", self.setOperatorReviewerDialog)
        fileMenu.addSeparator()
        fileMenu.addAction("Quit",self.quit,"Ctrl+Q")
        # This is a very bad way to do this, but I haven't worked anything else out (setMenuRole() didn't work)
        # Add it a second time, then it appears!
        if platform.system() == 'Darwin':
            fileMenu.addAction("&Quit",self.quit,"Ctrl+Q")

        specMenu = self.menuBar().addMenu("&Interface")

        self.useAmplitudeTick = specMenu.addAction("Show amplitude plot", self.useAmplitudeCheck)
        self.useAmplitudeTick.setCheckable(True)
        self.useAmplitudeTick.setChecked(self.config['showAmplitudePlot'])
        self.useAmplitude = True

        self.useFilesTick = specMenu.addAction("Show list of files", self.useFilesCheck)
        self.useFilesTick.setCheckable(True)
        self.useFilesTick.setChecked(self.config['showListofFiles'])

        self.showOverviewSegsTick = specMenu.addAction("Show annotation overview", self.showOverviewSegsCheck)
        self.showOverviewSegsTick.setCheckable(True)
        self.showOverviewSegsTick.setChecked(self.config['showAnnotationOverview'])

        self.dragRectangles = specMenu.addAction("Drag boxes in spectrogram", self.dragRectanglesCheck)
        self.dragRectangles.setCheckable(True)
        self.dragRectangles.setChecked(self.config['dragBoxes'])

        self.dragRectTransparent = specMenu.addAction("Make dragged boxes transparent", self.dragRectsTransparent)
        self.dragRectTransparent.setCheckable(True)
        self.dragRectTransparent.setChecked(self.config['transparentBoxes'])

        self.showFundamental = specMenu.addAction("Show fundamental frequency", self.showFundamentalFreq)
        self.showFundamental.setCheckable(True)
        self.showFundamental.setChecked(False)

        if self.DOC==False:
            self.showFundamental2 = specMenu.addAction("Show fundamental frequency2", self.showFundamentalFreq2)
            self.showFundamental2.setCheckable(True)
            self.showFundamental2.setChecked(False)

        if self.DOC==False:
            self.showInvSpec = specMenu.addAction("Show inverted spectrogram", self.showInvertedSpectrogram)
            self.showInvSpec.setCheckable(True)
            self.showInvSpec.setChecked(False)

        # if self.DOC==False:
        self.redoaxis = specMenu.addAction("Make frequency axis tight", self.redoFreqAxis)

        colMenu = specMenu.addMenu("&Choose colour map")
        colGroup = QActionGroup(self)
        for colour in self.config['ColourList']:
            cm = colMenu.addAction(colour)
            cm.setCheckable(True)
            if colour==self.config['cmap']:
                cm.setChecked(True)
            receiver = lambda cmap=colour: self.setColourMap(cmap)
            self.connect(cm, SIGNAL("triggered()"), receiver)
            colGroup.addAction(cm)
        self.invertcm = specMenu.addAction("Invert colour map",self.invertColourMap)
        self.invertcm.setCheckable(True)
        self.invertcm.setChecked(self.config['invertColourMap'])

        specMenu.addSeparator()
        specMenu.addAction("Change spectrogram parameters",self.showSpectrogramDialog)

        actionMenu = self.menuBar().addMenu("&Actions")
        actionMenu.addAction("&Delete all segments", self.deleteAll, "Ctrl+D")
        self.readonly = actionMenu.addAction("Make read only",self.makeReadOnly)
        self.readonly.setCheckable(True)
        self.readonly.setChecked(False)
        actionMenu.addAction("Denoise",self.denoiseDialog)
        actionMenu.addAction("Segment",self.segmentationDialog)
        actionMenu.addAction("Classify segments",self.classifySegments)
        actionMenu.addAction("Find matches",self.findMatches)
        if self.DOC==False:
            actionMenu.addAction("Filter spectrogram",self.medianFilterSpec)
            actionMenu.addAction("Denoise spectrogram",self.denoiseImage)
        actionMenu.addSeparator()
        actionMenu.addAction("Check segments [All segments]",self.humanClassifyDialog1)
        actionMenu.addAction("Check segments [Choose species]",self.humanClassifyDialog2)
        actionMenu.addSeparator()
        actionMenu.addAction("Put docks back",self.dockReplace)

        helpMenu = self.menuBar().addMenu("&Help")
        #aboutAction = QAction("About")
        helpMenu.addAction("About",self.showAbout)
        if platform.system() == 'Darwin':
            helpMenu.addAction("About",self.showAbout)
        helpMenu.addAction("Help",self.showHelp)
        helpMenu.addAction("Cheat Sheet", self.showCheatSheet)

    def showAbout(self):
        """ Create the About Message Box"""
        msg = QMessageBox()
        msg.setIconPixmap(QPixmap("img\AviaNZ.png"))
        msg.setWindowIcon(QIcon('img/Avianz.ico'))
        msg.setText("The AviaNZ Program, v0.10 (June 2017)")
        msg.setInformativeText("By Stephen Marsland, Massey University (2016--2017). With input from Nirosha Priyadarshani, Isabel Castro, Moira Pryde, Stuart Cockburn, Rebecca Stirnemann, Sumudu Manic Purage. \ns.r.marsland@massey.ac.nz; n.p.priyadarshani@massey.ac.nz")
        msg.setWindowTitle("About")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()
        return

    def showHelp(self):
        """ Show the user manual (a pdf file)"""
        import webbrowser
        webbrowser.open_new(r'file://' + os.path.realpath('./Docs/AviaNZManual.pdf'))

    def showCheatSheet(self):
        """ Show the cheat sheet of sample spectrograms (a pdf file)"""
        import webbrowser
        webbrowser.open_new(r'file://' + os.path.realpath('./Docs/CheatSheet.pdf'))

    def genConfigFile(self):
        """ If the configuration does exists, this generates one with default values for parameters. """
        print("Generating new config file")
        self.config = {
            'username': "Stephen",
            # Params for spectrogram
            'window_width': 256,
            'incr': 128,

            # Params for denoising
            'maxSearchDepth': 20,

            # Params for segmentation
            'minSegment': 50,
            'dirpath': './Sound Files',
            'secsSave': 60,

            # Param for width in seconds of the main representation
            'windowWidth': 10.0,

            # Text offset for labels
            'textoffset': 9,

            # Width of the segment markers in the overview plot (in seconds)
            'widthOverviewSegment': 10.0,

            # Max length of file to load at one time (in seconds), and overlap with next file
            'maxFileShow': 300,
            'fileOverlap': 10,

            # These are the contrast parameters for the spectrogram
            #'colourStart': 0.25,
            #'colourEnd': 0.75,
            'brightness': 50,
            'contrast': 50,

            # Params for cross-correlation and related
            'corrThr': 0.4,
            # Amount of overlap for 2 segments to be counted as the same
            # TODO: use this?
            'overlap_allowed': 5,

            #'BirdButtons1': ["Bellbird", "Bittern", "Cuckoo", "Fantail", "Hihi", "Kakapo", "Kereru", "Kiwi (F)", "Kiwi (M)",
            #                 "Petrel"],
            #'BirdButtons2': ["Rifleman", "Ruru", "Saddleback", "Silvereye", "Tomtit", "Tui", "Warbler", "Not Bird",
            #                 "Don't Know", "Other"],
            #'ListBirdsEntries': ['Albatross', 'Avocet', 'Blackbird', 'Bunting', 'Chaffinch', 'Egret', 'Gannet', 'Godwit',
            #                     'Gull', 'Kahu', 'Kaka', 'Kea', 'Kingfisher', 'Kokako', 'Lark', 'Magpie', 'Plover',
            #                     'Pukeko', "Rooster" 'Rook', 'Thrush', 'Warbler', 'Whio'],
            'BirdList': ["Bellbird", "Bittern", "Cuckoo", "Fantail", "Hihi", "Kakapo", "Kereru", "Kiwi (F)", "Kiwi (M)", "Kiwi", "Petrel","Rifleman", "Ruru", "Saddleback", "Silvereye", "Tomtit", "Tui", "Warbler", "Not Bird", "Don't Know",'Albatross', 'Avocet', 'Blackbird', 'Bunting', 'Chaffinch', 'Egret', 'Gannet', 'Godwit','Gull', 'Kahu', 'Kaka', 'Kea', 'Kingfisher', 'Kokako', 'Lark', 'Magpie', 'Plover','Pukeko', "Rooster" 'Rook', 'Thrush', 'Warbler', 'Whio'],

            'ColourList': ['Grey','Viridis', 'Inferno', 'Plasma', 'Autumn', 'Cool', 'Bone', 'Copper', 'Hot', 'Jet','Thermal','Flame','Yellowy','Bipolar','Spectrum'],
            # The colours for the segment boxes
            'ColourNone': (0, 0, 255, 100), # Blue
            'ColourSelected': (0, 255, 0, 100), # Green
            'ColourNamed': (255, 0, 0, 100), # Red
            'ColourPossible': (255, 255, 0, 100), # Yellow

            'cmap': 'Grey',

            # User interface paramaters
            'showAmplitudePlot': True,
            'showAnnotationOverview': True,
            'dragBoxes': False,
            'transparentBoxes': False,
            'showListofFiles': True,
            'invertColourMap': False
        }

    def createFrame(self):
        """ Creates the main window.
        This consists of a set of pyqtgraph docks with widgets in.
         d_ for docks, w_ for widgets, p_ for plots"""

        # Make the window and set its size
        self.area = DockArea()
        self.setCentralWidget(self.area)
        self.resize(1240,600)
        self.move(100,50)

        # Make the docks and lay them out
        self.d_overview = Dock("Overview",size = (1200,150))
        self.d_ampl = Dock("Amplitude",size=(1200,150))
        self.d_spec = Dock("Spectrogram",size=(1200,300))
        self.d_controls = Dock("Controls",size=(40,100))
        self.d_files = Dock("Files",size=(40,200))

        self.area.addDock(self.d_files,'left')
        self.area.addDock(self.d_overview,'right',self.d_files)
        self.area.addDock(self.d_ampl,'bottom',self.d_overview)
        self.area.addDock(self.d_spec,'bottom',self.d_ampl)
        self.area.addDock(self.d_controls,'bottom',self.d_files)

        # Put content widgets in the docks
        self.w_overview = pg.LayoutWidget()
        self.d_overview.addWidget(self.w_overview)
        self.w_overview1 = pg.GraphicsLayoutWidget()
        self.w_overview1.ci.layout.setContentsMargins(0.5, 1, 0.5, 1)
        self.w_overview.addWidget(self.w_overview1,row=0, col=2,rowspan=3)

        self.p_overview = self.w_overview1.addViewBox(enableMouse=False,enableMenu=False,row=0,col=0)
        self.p_overview2 = SupportClasses.ChildInfoViewBox(enableMouse=False, enableMenu=False)
        self.w_overview1.addItem(self.p_overview2,row=1,col=0)
        self.p_overview2.setXLink(self.p_overview)

        self.w_ampl = pg.GraphicsLayoutWidget()
        self.p_ampl = self.w_ampl.addViewBox(enableMouse=False,enableMenu=False)
        self.w_ampl.addItem(self.p_ampl,row=0,col=1)
        self.d_ampl.addWidget(self.w_ampl)

        self.w_spec = pg.GraphicsLayoutWidget()
        self.p_spec = SupportClasses.DragViewBox(enableMouse=False,enableMenu=False)
        self.w_spec.addItem(self.p_spec,row=0,col=1)
        self.d_spec.addWidget(self.w_spec)

        # The axes
        self.timeaxis = SupportClasses.TimeAxis(orientation='bottom')
        self.w_spec.addItem(self.timeaxis,row=1,col=1)
        self.timeaxis.linkToView(self.p_ampl)
        self.timeaxis.setLabel('Time',units='mm:ss')

        self.ampaxis = pg.AxisItem(orientation='left')
        self.w_ampl.addItem(self.ampaxis,row=0,col=0)
        self.ampaxis.linkToView(self.p_ampl)
        self.ampaxis.setWidth(w=65)
        self.ampaxis.setLabel('')

        self.specaxis = pg.AxisItem(orientation='left')
        self.w_spec.addItem(self.specaxis,row=0,col=0)
        self.specaxis.linkToView(self.p_spec)
        self.specaxis.setWidth(w=65)

        # The print out at the bottom of the spectrogram with data in
        self.pointData = pg.TextItem(color=(255,0,0),anchor=(0,0))
        self.p_spec.addItem(self.pointData)

        # The various plots
        self.overviewImage = pg.ImageItem(enableMouse=False)
        self.p_overview.addItem(self.overviewImage)
        self.amplPlot = pg.PlotDataItem()
        self.p_ampl.addItem(self.amplPlot)
        self.specPlot = pg.ImageItem()
        self.p_spec.addItem(self.specPlot)

        # Connect up the listeners
        # Have to connect up both of the spectogram ones so that one can be disconnected in the drag menu item listener
        self.p_ampl.scene().sigMouseClicked.connect(self.mouseClicked_ampl)
        #self.p_spec.sigMouseDragged.connect(self.mouseDragged_spec)
        #self.p_spec.scene().sigMouseClicked.connect(self.mouseClicked_spec)
        self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)

        # The content of the other two docks
        self.w_controls = pg.LayoutWidget()
        self.d_controls.addWidget(self.w_controls)
        self.w_files = pg.LayoutWidget()
        self.d_files.addWidget(self.w_files)

        # The buttons to move through the overview
        self.leftBtn = QToolButton()
        self.leftBtn.setArrowType(Qt.LeftArrow)
        self.connect(self.leftBtn, SIGNAL('clicked()'), self.moveLeft)
        self.w_overview.addWidget(self.leftBtn,row=0,col=0)
        self.rightBtn = QToolButton()
        self.rightBtn.setArrowType(Qt.RightArrow)
        self.connect(self.rightBtn, SIGNAL('clicked()'), self.moveRight)
        self.w_overview.addWidget(self.rightBtn,row=0,col=1)

        # Button to move to the next file in the list
        self.nextFileBtn=QToolButton()
        self.nextFileBtn.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaSkipForward))
        self.connect(self.nextFileBtn, SIGNAL('clicked()'), self.openNextFile)
        self.w_files.addWidget(self.nextFileBtn,row=0,col=1)
        #self.w_overview.addWidget(self.nextFileBtn,row=1,colspan=2)

        # Buttons to move to next/previous five minutes
        self.prev5mins=QToolButton()
        self.prev5mins.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaSeekBackward))
        self.connect(self.prev5mins, SIGNAL('clicked()'), self.movePrev5mins)
        self.w_overview.addWidget(self.prev5mins,row=1,col=0)
        self.next5mins=QToolButton()
        self.next5mins.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaSeekForward))
        self.connect(self.next5mins, SIGNAL('clicked()'), self.moveNext5mins)
        self.w_overview.addWidget(self.next5mins,row=1,col=1)
        # TODO: Add a label -- how to squeeze it into the space?
        self.placeInFileLabel = QLabel('')
        self.w_overview.addWidget(self.placeInFileLabel,row=2,colspan=2)

        # The buttons inside the controls dock
        self.playButton = QtGui.QToolButton()
        self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))
        self.playButton.setToolTip("Play visible")
        self.connect(self.playButton, SIGNAL('clicked()'), self.playSegment)

        self.playSegButton = QtGui.QToolButton()
        self.playSegButton.setIcon(QtGui.QIcon('img/playsegment.png'))
        self.playSegButton.setIconSize(QtCore.QSize(20, 20))
        self.playSegButton.setToolTip("Play selected")
        self.connect(self.playSegButton, SIGNAL('clicked()'), self.playSelectedSegment)
        self.playSegButton.setEnabled(False)

        self.playBandLimitedSegButton = QtGui.QToolButton()
        self.playBandLimitedSegButton.setIcon(QtGui.QIcon('img/playBandLimited.png'))
        self.playBandLimitedSegButton.setToolTip("Play selected-band limited")
        self.playBandLimitedSegButton.setIconSize(QtCore.QSize(20, 20))
        self.connect(self.playBandLimitedSegButton, SIGNAL('clicked()'), self.playBandLimitedSegment)
        self.playBandLimitedSegButton.setEnabled(False)

        # The slider to show playback position
        # This is hidden, but controls the moving bar
        self.playSlider = QSlider(Qt.Horizontal)
        self.connect(self.playSlider,SIGNAL('sliderReleased()'),self.sliderMoved)
        self.playSlider.setVisible(False)
        self.d_spec.addWidget(self.playSlider)
        self.timePlayed = QLabel()

        # A slider to move through the file easily
        self.scrollSlider = QScrollBar(Qt.Horizontal)
        self.scrollSlider.valueChanged.connect(self.scroll)
        self.d_spec.addWidget(self.scrollSlider)

        # The spinbox for changing the width shown in the controls dock
        self.widthWindow = QDoubleSpinBox()
        self.widthWindow.setSingleStep(1.0)
        self.widthWindow.setDecimals(2)
        self.widthWindow.setValue(self.config['windowWidth'])
        self.widthWindow.valueChanged[float].connect(self.changeWidth)

        # Brightness, and contrast sliders
        self.brightnessSlider = QSlider(Qt.Horizontal)
        self.brightnessSlider.setMinimum(0)
        self.brightnessSlider.setMaximum(100)
        self.brightnessSlider.setValue(self.config['brightness'])
        self.brightnessSlider.setTickInterval(1)
        self.brightnessSlider.valueChanged.connect(self.setColourLevels)

        self.contrastSlider = QSlider(Qt.Horizontal)
        self.contrastSlider.setMinimum(0)
        self.contrastSlider.setMaximum(100)
        self.contrastSlider.setValue(self.config['contrast'])
        self.contrastSlider.setTickInterval(1)
        self.contrastSlider.valueChanged.connect(self.setColourLevels)

        # Delete segment button
        deleteButton = QPushButton("&Delete Current Segment")
        self.connect(deleteButton, SIGNAL('clicked()'), self.deleteSegment)

        # Place all these widgets in the Controls dock
        self.w_controls.addWidget(self.playButton,row=0,col=0)
        self.w_controls.addWidget(self.playSegButton,row=0,col=1)
        self.w_controls.addWidget(self.playBandLimitedSegButton,row=0,col=2)
        self.w_controls.addWidget(self.timePlayed,row=1,col=0)
        self.w_controls.addWidget(QLabel("Brightness"),row=2,col=0,colspan=3)
        self.w_controls.addWidget(self.brightnessSlider,row=3,col=0,colspan=3)
        self.w_controls.addWidget(QLabel("Contrast"),row=4,col=0,colspan=3)
        self.w_controls.addWidget(self.contrastSlider,row=5,col=0,colspan=3)
        self.w_controls.addWidget(deleteButton,row=6,col=0,colspan=3)
        self.w_controls.addWidget(QLabel('Visible window width (seconds)'),row=7,col=0,colspan=3)
        self.w_controls.addWidget(self.widthWindow,row=8,col=0,colspan=3)#,colspan=2)

        # List to hold the list of files
        self.listFiles = QListWidget(self)
        self.listFiles.setMinimumWidth(150)
        self.listFiles.connect(self.listFiles, SIGNAL('itemDoubleClicked(QListWidgetItem*)'), self.listLoadFile)

        self.w_files.addWidget(QLabel('Double click to open'),row=0,col=0)
        self.w_files.addWidget(QLabel('Red names have segments'),row=1,col=0)
        self.w_files.addWidget(self.listFiles,row=2,colspan=2)

        # The context menu (drops down on mouse click) to select birds
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.menuBirdList = QMenu()
        self.menuBird2 = self.menuBirdList.addMenu('Other')
        self.fillBirdList()

        # Audio playback
        # Instantiate a Qt media object and prepare it
        self.media_obj = phonon.Phonon.MediaObject(self)
        self.audio_output = phonon.Phonon.AudioOutput(phonon.Phonon.MusicCategory, self)
        phonon.Phonon.createPath(self.media_obj, self.audio_output)
        self.media_obj.setTickInterval(20)
        self.media_obj.finished.connect(self.playFinished)
        self.media_obj.tick.connect(self.movePlaySlider)

        self.volSlider = phonon.Phonon.VolumeSlider()
        self.volSlider.setOrientation(Qt.Horizontal)
        self.volSlider.setGeometry(QtCore.QRect(50, 50, 150, 40))
        self.volSlider.setFixedWidth(150)
        self.volSlider.setMaximumVolume(1.0)
        self.volSlider.setAudioOutput(self.audio_output)
        self.w_controls.addWidget(self.volSlider,row=1,col=1,colspan=2)

        # Make the colours that are used in the interface
        # The dark ones are to draw lines instead of boxes
        self.ColourSelected = QtGui.QColor(self.config['ColourSelected'][0], self.config['ColourSelected'][1], self.config['ColourSelected'][2], self.config['ColourSelected'][3])
        self.ColourNamed = QtGui.QColor(self.config['ColourNamed'][0], self.config['ColourNamed'][1], self.config['ColourNamed'][2], self.config['ColourNamed'][3])
        self.ColourNone = QtGui.QColor(self.config['ColourNone'][0], self.config['ColourNone'][1], self.config['ColourNone'][2], self.config['ColourNone'][3])
        self.ColourPossible = QtGui.QColor(self.config['ColourPossible'][0], self.config['ColourPossible'][1], self.config['ColourPossible'][2], self.config['ColourPossible'][3])

        self.ColourSelectedDark = QtGui.QColor(self.config['ColourSelected'][0], self.config['ColourSelected'][1], self.config['ColourSelected'][2], 255)
        self.ColourNamedDark = QtGui.QColor(self.config['ColourNamed'][0], self.config['ColourNamed'][1], self.config['ColourNamed'][2], 255)
        self.ColourNoneDark = QtGui.QColor(self.config['ColourNone'][0], self.config['ColourNone'][1], self.config['ColourNone'][2], 255)
        self.ColourPossibleDark = QtGui.QColor(self.config['ColourPossible'][0], self.config['ColourPossible'][1], self.config['ColourPossible'][2], 255)

        # Hack to get the type of an ROI
        p_spec_r = SupportClasses.ShadedRectROI(0, 0)
        self.ROItype = type(p_spec_r)

        # Listener for key presses
        self.connect(self.p_spec, SIGNAL("keyPressed"),self.handleKey)

        # Store the state of the docks in case the user wants to reset it
        self.state = self.area.saveState()

        # Function calls to check if should show various parts of the interface, whether dragging boxes or not
        self.useAmplitudeCheck()
        self.useFilesCheck()
        self.showOverviewSegsCheck()
        self.dragRectsTransparent()
        self.dragRectanglesCheck()

        # add statusbar
        self.statusLeft = QLabel("Left")
        self.statusLeft.setFrameStyle(QFrame.Panel) #,QFrame.Sunken)
        self.statusMid = QLabel("????")
        self.statusMid.setFrameStyle(QFrame.Panel) #,QFrame.Sunken)
        self.statusRight = QLabel("Right")
        self.statusRight.setAlignment(Qt.AlignRight)
        self.statusRight.setFrameStyle(QFrame.Panel) #,QFrame.Sunken)
        # Style
        statusStyle='QLabel {border:transparent}'
        self.statusLeft.setStyleSheet(statusStyle)
        # self.statusMid.setStyleSheet(statusStyle)
        self.statusRight.setStyleSheet(statusStyle)
        self.statusBar().addPermanentWidget(self.statusLeft,1)
        # self.statusBar().addPermanentWidget(self.statusMid,1)
        self.statusBar().addPermanentWidget(self.statusRight,1)

        # Set the message in the status bar
        self.statusLeft.setText("Ready")
        self.statusRight.setText("Operator/Reviewer:")

        # Plot everything
        self.show()

    def keyPressEvent(self,ev):
        """ Listener to handle keypresses and emit a keypress event, which is dealt with by handleKey()"""
        self.emit(SIGNAL("keyPressed"),ev)

    def handleKey(self,ev):
        """ Handle keys pressed during program use.
        These are:
            backspace to delete a segment
            escape to pause playback """
        if ev.key() == Qt.Key_Backspace:
            self.deleteSegment()
        elif ev.key()==Qt.Key_Escape:
            if self.media_obj.state() != phonon.Phonon.PausedState or self.media_obj.state() != phonon.Phonon.StoppedState:
                self.media_obj.pause()
                self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))

    def fillBirdList(self,unsure=False):
        """ Sets the contents of the context menu.
        The first 20 items are in the first menu, the next in a second menu.
        This is called a lot because the order of birds in the list changes since the last choice
        is moved to the top of the list. """
        self.menuBirdList.clear()
        self.menuBird2.clear()
        for item in self.config['BirdList'][:20]:
            if unsure and item != "Don't Know":
                item = item+'?'
            bird = self.menuBirdList.addAction(item)
            receiver = lambda birdname=item: self.birdSelected(birdname)
            self.connect(bird, SIGNAL("triggered()"), receiver)
            self.menuBirdList.addAction(bird)
        self.menuBird2 = self.menuBirdList.addMenu('Other')
        for item in self.config['BirdList'][20:]+['Other']:
            if unsure and item != "Don't Know" and item != "Other":
                item = item+'?'
            bird = self.menuBird2.addAction(item)
            receiver = lambda birdname=item: self.birdSelected(birdname)
            self.connect(bird, SIGNAL("triggered()"), receiver)
            self.menuBird2.addAction(bird)

    def fillFileList(self,fileName):
        """ Generates the list of files for the file listbox.
        Most of the work is to deal with directories in that list.
        It only sees *.wav files. Picks up *.data and *_1.wav files, the first to make the filenames
        red in the list, and the second to know if the files are long."""
        if not os.path.isdir(self.dirName):
            print("Directory doesn't exist: making it")
            os.makedirs(self.dirName)

        self.listOfFiles = QDir(self.dirName).entryInfoList(['..','*.wav'],filters=QDir.AllDirs|QDir.NoDot|QDir.Files,sort=QDir.DirsFirst)
        listOfDataFiles = QDir(self.dirName).entryList(['*.data'])
        listOfLongFiles = QDir(self.dirName).entryList(['*_1.wav'])
        for file in self.listOfFiles:
            if file.fileName()[:-4]+'_1.wav' in listOfLongFiles:
                # Ignore this entry
                pass
            else:
                # If there is a .data version, colour the name red to show it has been labelled
                item = QListWidgetItem(self.listFiles)
                self.listitemtype = type(item)
                item.setText(file.fileName())
                if file.fileName()+'.data' in listOfDataFiles:
                    item.setTextColor(Qt.red)
        if fileName:
            index = self.listFiles.findItems(fileName,Qt.MatchExactly)
            if len(index)>0:
                self.listFiles.setCurrentItem(index[0])
            else:
                index = self.listFiles.findItems(self.listOfFiles[0].fileName(),Qt.MatchExactly)
                self.listFiles.setCurrentItem(index[0])

    def resetStorageArrays(self):
        """ Called when new files are loaded.
        Resets the variables that hold the data to be saved and/or plotted. """

        # Remove the segments
        self.removeSegments()
        if hasattr(self, 'overviewImageRegion'):
            self.p_overview.removeItem(self.overviewImageRegion)

        # This is a flag to say if the next thing that the user clicks on should be a start or a stop for segmentation
        if self.started:
            # This is the second click, so should pay attention and close the segment
            # Stop the mouse motion connection, remove the drawing boxes
            if self.started_window=='a':
                self.p_ampl.scene().sigMouseMoved.disconnect()
                self.p_ampl.removeItem(self.vLine_a)
            else:
                self.p_spec.scene().sigMouseMoved.disconnect()
                # Add the other mouse move listener back
                self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
                self.p_spec.removeItem(self.vLine_s)
            self.p_ampl.removeItem(self.drawingBox_ampl)
            self.p_spec.removeItem(self.drawingBox_spec)
        self.started = False

        # Keep track of start points and selected buttons
        self.windowStart = 0
        self.playPosition = self.windowStart
        self.prevBoxCol = self.config['ColourNone']
        self.bar.setValue(0)

        # Delete the overview segments
        for r in self.SegmentRects:
            self.p_overview2.removeItem(r)
        self.SegmentRects = []

        # Remove any fundamental frequencies drawn
        for r in self.segmentPlots:
            self.p_spec.removeItem(r)
        self.segmentPlots=[]

        #self.nFileSections = None

    def openFile(self):
        """ This handles the menu item for opening a file.
        Splits the directory name and filename out, and then passes the filename to the loader."""
        fileName = QtGui.QFileDialog.getOpenFileName(self, 'Choose File', self.dirName,"Wav files (*.wav)")

        # Find the '/' in the fileName
        i=len(fileName)-1
        while fileName[i] != '/' and i>0:
            i = i-1
        self.dirName = fileName[:i+1]

        self.listLoadFile(fileName)

    def listLoadFile(self,current):
        """ Listener for when the user clicks on a filename (also called by openFile() )
        Prepares the program for a new file.
        Saves the segments of the current file, resets flags and calls loadFile() """

        # If there was a previous file, make sure the type of its name is OK. This is because you can get these
        # names from the file listwidget, or from the openFile dialog.
        if self.previousFile is not None:
            if type(self.previousFile) is not self.listitemtype:
                self.previousFile = self.listFiles.findItems(os.path.basename(str(self.previousFile)), Qt.MatchExactly)
                if len(self.previousFile)>0:
                    self.previousFile = self.previousFile[0]

            if self.segments != [] or self.hasSegments:
                if len(self.segments)>0:
                    if self.segments[0][0] > -1:
                        self.segments.insert(0, [-1, -1, str(self.username), -1, -1])
                else:
                    self.segments.insert(0, [-1, -1, str(self.username), -1, -1])
                self.saveSegments()
                self.previousFile.setTextColor(Qt.red)
        self.previousFile = current
        self.resetStorageArrays()

        # Reset the media player
        if self.media_obj.state() == phonon.Phonon.PlayingState:
            self.media_obj.pause()
            self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))

        if type(current) is QString:
            pass
        else:
            current = current.text()

        # Update the file list to show the right one
        i=0
        while self.listOfFiles[i].fileName() != current and i<len(self.listOfFiles)-1:
            i+=1
        if self.listOfFiles[i].isDir() or (i == len(self.listOfFiles)-1 and self.listOfFiles[i].fileName() != current):
            dir = QDir(self.dirName)
            dir.cd(self.listOfFiles[i].fileName())
            # Now repopulate the listbox
            self.dirName=str(dir.absolutePath())
            self.listFiles.clearSelection()
            self.listFiles.clearFocus()
            self.listFiles.clear()
            self.previousFile = None
            if (i == len(self.listOfFiles)-1) and (self.listOfFiles[i].fileName() != current):
                self.loadFile(current)
            self.fillFileList(current)
            # Show the selected file
            index = self.listFiles.findItems(os.path.basename(str(current)), Qt.MatchExactly)
            if len(index) > 0:
                self.listFiles.setCurrentItem(index[0])
        else:
            self.loadFile(current)

    def loadFile(self,name=None):
        """ This does the work of loading a file.
        We are using wavio to do the reading. We turn the data into a float, but do not normalise it (/2^(15)).
        For 2 channels, just take the first one.
        Normalisation can cause problems for some segmentations, e.g. Harma.

        If no name is specified, loads the next section of the current file

        This method also gets the spectrogram to plot it, loads the segments from a *.data file, and
        passes the new data to any of the other classes that need it.
        Then it sets up the audio player and fills in the appropriate time data in the window, and makes
        the scroll bar and overview the appropriate lengths.
        """

        with pg.ProgressDialog("Loading..", 0, 7) as dlg:
            dlg.setCancelButton(None)
            if name is not None:
                if isinstance(name,str):
                    self.filename = self.dirName+'/'+name
                elif isinstance(name,QString):
                    name = os.path.basename(str(name))
                    self.filename = self.dirName+'/'+ name
                else:
                    self.filename = str(self.dirName+'/'+name.text())
                dlg += 1

                # Create an instance of the Signal Processing class
                if not hasattr(self, 'sp'):
                    self.sp = SignalProc.SignalProc([],0,self.config['window_width'],self.config['incr'])

                self.currentFileSection = 0
                self.timeaxis.setOffset(0)

                dlg += 1
            else:
                dlg += 2

            # Read in the file and make the spectrogram
            self.startRead = max(0,self.currentFileSection*self.config['maxFileShow']-self.config['fileOverlap'])
            if self.startRead == 0:
                self.lenRead = self.config['maxFileShow']+self.config['fileOverlap']
            else:
                self.lenRead = self.config['maxFileShow'] + 2*self.config['fileOverlap']

            wavobj = wavio.read(self.filename,self.lenRead,self.startRead)
            #wavobj = wavio.read(self.filename)
            self.sampleRate = wavobj.rate
            self.audiodata = wavobj.data
            self.minFreq = 0
            self.maxFreq = self.sampleRate / 2.
            self.fileLength = wavobj.nframes
            self.timeaxis.setOffset(self.startRead)
            dlg += 1

            if self.audiodata.dtype is not 'float':
                self.audiodata = self.audiodata.astype('float')  # / 32768.0

            if np.shape(np.shape(self.audiodata))[0] > 1:
                self.audiodata = self.audiodata[:, 0]
            self.datalength = np.shape(self.audiodata)[0]
            print("Length of file is ", self.datalength, float(self.datalength) / self.sampleRate), "loaded from ", self.fileLength, float(self.fileLength) / self.sampleRate

            if name is not None:
                if self.datalength != self.fileLength:
                    print "not all of file loaded"
                    self.nFileSections = int(np.ceil(float(self.fileLength)/self.datalength))
                    print self.nFileSections, self.currentFileSection
                    self.prev5mins.setEnabled(False)
                    self.next5mins.setEnabled(True)
                    #self.placeInFileLabel.setText('')
                else:
                    self.nFileSections = None
                    self.prev5mins.setEnabled(False)
                    self.next5mins.setEnabled(False)
                    # self.placeInFileLabel.setText('')

            if self.nFileSections is None:
                self.placeInFileLabel.setText('')
            else:
                self.placeInFileLabel.setText("Part "+ str(self.currentFileSection+1) + " of " + str(self.nFileSections))

            # Get the data for the main spectrogram
            sgRaw = self.sp.spectrogram(self.audiodata, self.sampleRate, self.config['window_width'],
                                        self.config['incr'], mean_normalise=True, onesided=True,
                                        multitaper=False)
            maxsg = np.min(sgRaw)
            self.sg = np.abs(np.where(sgRaw == 0, 0.0, 10.0 * np.log10(sgRaw / maxsg)))

            # Load any previous segments stored
            if os.path.isfile(self.filename + '.data'):
                file = open(self.filename + '.data', 'r')
                self.segments = json.load(file)
                file.close()
                if len(self.segments) > 0:
                    if self.segments[0][0] == -1:
                        del self.segments[0]
                self.hasSegments = True
            else:
                self.hasSegments = False

            # Update the data that is seen by the other classes
            if hasattr(self,'seg'):
                self.seg.setNewData(self.audiodata,sgRaw,self.sampleRate,self.config['window_width'],self.config['incr'])
            else:
                self.seg = Segment.Segment(self.audiodata, sgRaw, self.sp, self.sampleRate, self.config['minSegment'],
                                           self.config['window_width'], self.config['incr'])
            self.sp.setNewData(self.audiodata,self.sampleRate)

            # Delete any denoising backups from the previous file
            if hasattr(self,'audiodata_backup'):
                self.audiodata_backup = None
            self.showFundamental.setChecked(False)
            if self.DOC == False:
                self.showInvSpec.setChecked(False)

            # Set the window size
            self.windowSize = self.config['windowWidth']
            self.widthWindow.setRange(0.5, float(len(self.audiodata))/self.sampleRate)

            # Reset it if the file is shorter than the window
            if float(len(self.audiodata))/self.sampleRate < self.windowSize:
                self.windowSize = float(len(self.audiodata))/self.sampleRate
            self.widthWindow.setValue(self.windowSize)

            self.totalTime = self.convertMillisecs((float(self.datalength)/self.sampleRate)*1000)

            # Load the file for playback as well, and connect up the listeners for it
            self.media_obj.setCurrentSource(phonon.Phonon.MediaSource(self.filename))

            # Set the length of the scrollbar **** Changed to try to match full size of file
            #self.scrollSlider.setRange(0,np.shape(self.sg)[0]-self.convertAmpltoSpec(self.widthWindow.value()))
            self.scrollSlider.setRange(0,self.convertAmpltoSpec(self.fileLength)-self.convertAmpltoSpec(self.widthWindow.value()))
            self.scrollSlider.setValue(0)

            # Get the height of the amplitude for plotting the box
            self.minampl = np.min(self.audiodata)+0.1*(np.max(self.audiodata)+np.abs(np.min(self.audiodata)))
            self.drawOverview()
            dlg += 1
            self.drawfigMain()
            self.setWindowTitle('AviaNZ - ' + self.filename)
            dlg += 1
            self.statusLeft.setText("Ready")

    def openNextFile(self):
        """ Listener for next file >> button.
        Get the next file in the list and call the loader. """
        i=self.listFiles.currentRow()
        if i+1<len(self.listFiles):
            self.listFiles.setCurrentRow(i+1)
            self.listLoadFile(self.listFiles.currentItem())
        else:
            # Tell the user they've finished
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("You've finished processing the folder")
            msg.setWindowIcon(QIcon('img/Avianz.ico'))
            msg.setWindowTitle("Last file")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    def dragRectanglesCheck(self):
        """ Listener for the menuitem that says if the user is dragging rectangles or clicking on the spectrogram has
        changed state.
        Changes the pyqtgraph MouseMode.
        Also swaps the listeners"""
        if self.dragRectangles.isChecked():
            self.p_spec.setMouseMode(pg.ViewBox.RectMode)
            try:
                self.p_spec.scene().sigMouseClicked.disconnect()
            except Exception:
                pass
            self.p_spec.sigMouseDragged.connect(self.mouseDragged_spec)
        else:
            self.p_spec.setMouseMode(pg.ViewBox.PanMode)
            try:
                self.p_spec.sigMouseDragged.disconnect()
            except Exception:
                pass
            self.p_spec.scene().sigMouseClicked.connect(self.mouseClicked_spec)

        self.config['dragBoxes'] = self.dragRectangles.isChecked()

    def dragRectsTransparent(self):
        """ Listener for the check menu item that decides if the user wants the dragged rectangles to have colour or not.
        It's a switch from Brush to Pen or vice versa.
        """
        if self.dragRectTransparent.isChecked():
            for box in self.listRectanglesa2:
                if type(box) == self.ROItype:
                    col = box.brush.color()
                    col.setAlpha(255)
                    box.setBrush(pg.mkBrush(None))
                    box.setPen(pg.mkPen(col,width=1))
                    box.update()
        else:
            for box in self.listRectanglesa2:
                if type(box) == self.ROItype:
                    col = box.pen.color()
                    col.setAlpha(self.ColourNamed.alpha())
                    box.setBrush(pg.mkBrush(col))
                    box.setPen(pg.mkPen(None))
                    box.update()
        self.config['transparentBoxes'] = self.dragRectTransparent.isChecked()

    def useAmplitudeCheck(self):
        """ Listener for the check menu item saying if the user wants to see the waveform.
        Does not remove the dock, just hide it. It's therefore easy to replace, but could have some performance overhead.
        """
        if self.useAmplitudeTick.isChecked():
            self.useAmplitude = True
            self.d_ampl.show()
        else:
            self.useAmplitude = False
            self.d_ampl.hide()
        self.config['showAmplitudePlot'] = self.useAmplitudeTick.isChecked()

    def useFilesCheck(self):
        """ Listener to process if the user swaps the check menu item to see the file list. """
        if self.useFilesTick.isChecked():
            self.d_files.show()
        else:
            self.d_files.hide()
        self.config['showListofFiles'] = self.useFilesTick.isChecked()

    def showOverviewSegsCheck(self):
        """ Listener to process if the user swaps the check menu item to see the overview segment boxes. """
        if self.showOverviewSegsTick.isChecked():
            self.p_overview2.show()
        else:
            self.p_overview2.hide()
        self.config['showAnnotationOverview'] = self.showOverviewSegsTick.isChecked()

    def makeReadOnly(self):
        """ Listener to process the check menu item to make the plots read only.
        Turns off the listeners for the amplitude and spectrogram plots.
        Also has to go through all of the segments, turn off the listeners, and make them unmovable.
        """
        if self.readonly.isChecked():
            self.p_ampl.scene().sigMouseClicked.disconnect()
            if self.dragRectangles.isChecked():
                self.p_spec.sigMouseDragged.disconnect()
            else:
                self.p_spec.scene().sigMouseClicked.disconnect()
            self.p_spec.scene().sigMouseMoved.disconnect()
            for rect in self.listRectanglesa1:
                rect.sigRegionChangeFinished.disconnect()
                rect.setMovable(False)
            for rect in self.listRectanglesa2:
                rect.sigRegionChangeFinished.disconnect()
                rect.setMovable(False)
        else:
            self.p_ampl.scene().sigMouseClicked.connect(self.mouseClicked_ampl)
            if self.dragRectangles.isChecked():
                self.p_spec.sigMouseDragged.connect(self.mouseDragged_spec)
            else:
                self.p_spec.scene().sigMouseClicked.connect(self.mouseClicked_spec)
            self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
            for rect in self.listRectanglesa1:
                rect.sigRegionChangeFinished.connect(self.updateRegion_ampl)
                rect.setMovable(True)
            for rect in self.listRectanglesa2:
                rect.sigRegionChangeFinished.connect(self.updateRegion_spec)
                rect.setMovable(True)

    def dockReplace(self):
        """ Listener for if the docks should be replaced menu item. """
        self.area.restoreState(self.state)

    def showFundamentalFreq(self):
        """ Calls the SignalProc class to compute, and then draws the fundamental frequency.
        Uses the yin algorithm. """
        if self.showFundamental.isChecked():
            self.statusLeft.setText("Drawing fundamental frequency...")
            pitch, y, minfreq, W = self.seg.yin()
            ind = np.squeeze(np.where(pitch>minfreq))
            pitch = pitch[ind]
            ind = ind*W/(self.config['window_width'])
            x = (pitch*2./self.sampleRate*np.shape(self.sg)[1]).astype('int')

            # TODO: Fit a spline and draw that instead
            #from scipy.interpolate import interp1d
            #f = interp1d(x, ind, kind='cubic')
            #self.sg[ind,x] = 1

            # Get the individual pieces
            segs = self.seg.identifySegments(ind,maxgap=10,minlength=5)
            count = 0
            self.segmentPlots = []
            for s in segs:
                count += 1
                s[0] = s[0] * self.sampleRate / float(self.config['incr'])
                s[1] = s[1] * self.sampleRate / float(self.config['incr'])
                i = np.where((ind>s[0]) & (ind<s[1]))
                self.segmentPlots.append(pg.PlotDataItem())
                self.segmentPlots[-1].setData(ind[i], x[i], pen=pg.mkPen('r', width=1))
                self.p_spec.addItem(self.segmentPlots[-1])
        else:
            self.statusLeft.setText("Removing fundamental frequency...")
            for r in self.segmentPlots:
                self.p_spec.removeItem(r)
        self.statusLeft.setText("Ready")

    def showFundamentalFreq2(self):
        # This and the next function are to check whether or not yaapt or harvest are any good. They aren't.
        import pYAAPT
        import basic_tools
        # Actually this is a pain, since it either gives back a value for each amplitude sample, or for it's own weird windows
        if self.showFundamental2.isChecked():
            y = basic_tools.SignalObj(self.filename)
            x = pYAAPT.yaapt(y)
            self.yinRois = []
            for r in range(len(x)):
                self.yinRois.append(pg.CircleROI([ind[r],x[r]], [2,2], pen=(4, 9),movable=False))
            for r in self.yinRois:
                self.p_spec.addItem(r)
        else:
            for r in self.yinRois:
                self.p_spec.removeItem(r)

    def showFundamentalFreq3(self):
        # Harvest
        import audio_tools
        if self.showFundamental2.isChecked():
            p, f, t, fa = audio_tools.harvest(self.audiodata,self.sampleRate)
            ind = f/self.config['window_width']
            x = (p*2./self.sampleRate*np.shape(self.sg)[1]).astype('int')

            self.yinRois = []
            for r in range(len(x)):
                self.yinRois.append(pg.CircleROI([ind[r],x[r]], [2,2], pen=(4, 9),movable=False))
            for r in self.yinRois:
                self.p_spec.addItem(r)
        else:
            for r in self.yinRois:
                self.p_spec.removeItem(r)

    def showInvertedSpectrogram(self):
        """ Listener for the menu item that draws the spectrogram of the waveform of the inverted spectrogram."""
        if self.showInvSpec.isChecked():
            sgRaw = self.sp.show_invS()
        else:
            sgRaw = self.sp.spectrogram(self.audiodata, self.sampleRate, mean_normalise=True, onesided=True,
                                         multitaper=False)
        maxsg = np.min(sgRaw)
        self.sg = np.abs(np.where(sgRaw == 0, 0.0, 10.0 * np.log10(sgRaw / maxsg)))
        self.overviewImage.setImage(self.sg)
        self.specPlot.setImage(self.sg)

    def medianFilterSpec(self):
        """ Median filter the spectrogram. To be used in conjunction with spectrogram inversion. """
        # TODO: Play with this
        self.statusLeft.setText("Filtering...")
        from scipy.ndimage.filters import median_filter
        median_filter(self.sg,size=(100,20))
        self.specPlot.setImage(self.sg)
        self.statusLeft.setText("Ready")

    def denoiseImage(self):
        """ Denoise the spectrogram. To be used in conjunction with spectrogram inversion. """
        from cv2 import fastNlMeansDenoising
        sg = np.array(self.sg/np.max(self.sg)*255,dtype = np.uint8)
        sg = fastNlMeansDenoising(sg,10,7,21)
        self.specPlot.setImage(sg)
# ==============
# Code for drawing and using the main figure

    def convertAmpltoSpec(self,x):
        """ Unit conversion """
        return x*self.sampleRate/self.config['incr']

    def convertSpectoAmpl(self,x):
        """ Unit conversion """
        return x*self.config['incr']/self.sampleRate

    def convertMillisecs(self,millisecs):
        """ Unit conversion """
        seconds = (millisecs / 1000) % 60
        minutes = (millisecs / (1000 * 60)) % 60
        return "%02d" % minutes+":"+"%02d" % seconds

    def drawOverview(self):
        """ On loading a new file, update the overview figure to show where you are up to in the file.
        Also, compute the new segments for the overview, and make sure that the listeners are connected
        for clicks on them. """
        self.overviewImage.setImage(self.sg)
        self.overviewImageRegion = pg.LinearRegionItem()
        self.p_overview.addItem(self.overviewImageRegion, ignoreBounds=True)
        self.overviewImageRegion.setRegion([0, self.convertAmpltoSpec(self.widthWindow.value())])
        print self.convertAmpltoSpec(self.widthWindow.value())

        self.overviewImageRegion.sigRegionChangeFinished.connect(self.updateOverview)

        # Three y values are No. not known, No. known, No. possible
        # widthOverviewSegment is in seconds
        numSegments = int(np.ceil(np.shape(self.sg)[0]/self.convertAmpltoSpec(self.config['widthOverviewSegment'])))
        self.widthOverviewSegment = int(float(np.shape(self.sg)[0])/numSegments)
        #print numSegments, self.widthOverviewSegment
        #numSegments = int(np.ceil(self.convertAmpltoSpec(float(self.fileLength)/self.sampleRate)/self.convertAmpltoSpec(self.config['widthOverviewSegment'])))
        #self.widthOverviewSegment = int(self.convertAmpltoSpec(float(self.fileLength)/self.sampleRate)/numSegments) -1
        #print numSegments, self.widthOverviewSegment, numSegments*self.widthOverviewSegment, self.config['widthOverviewSegment']

        self.overviewSegments = np.zeros((numSegments,3))
        for i in range(numSegments):
            r = SupportClasses.ClickableRectItem(i*self.widthOverviewSegment, 0, self.widthOverviewSegment, 0.5)
            r.setPen(pg.mkPen('k'))
            r.setBrush(pg.mkBrush('w'))
            self.SegmentRects.append(r)
            self.p_overview2.addItem(r)
        self.p_overview2.sigChildMessage.connect(self.overviewSegmentClicked)

    def overviewSegmentClicked(self,x):
        """ Listener for an overview segment being clicked on.
        Work out which one, and move the region appropriately. Calls updateOverview to do the work. """
        minX, maxX = self.overviewImageRegion.getRegion()
        self.overviewImageRegion.setRegion([x, x+maxX-minX])
        self.updateOverview()
        self.playPosition = int(self.convertSpectoAmpl(x)*1000.0)

    def updateOverview(self):
        """ Listener for when the overview box is changed. Also called by overviewSegmentClicked().
        Does the work of keeping all the plots in the right place as the overview moves.
        It sometimes updates a bit slowly. """
        # **** Will need work to deal with the paging

        minX, maxX = self.overviewImageRegion.getRegion()
        self.widthWindow.setValue(self.convertSpectoAmpl(maxX-minX))
        self.p_ampl.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), padding=0)
        self.p_spec.setXRange(minX, maxX, padding=0)
        self.p_ampl.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), padding=0)
        self.p_spec.setXRange(minX, maxX, padding=0)
        self.setSliderLimits(1000.0*self.convertSpectoAmpl(minX),1000.0*self.convertSpectoAmpl(maxX))
        self.scrollSlider.setValue(minX)
        self.pointData.setPos(minX,0)
        self.config['windowWidth'] = self.convertSpectoAmpl(maxX-minX)
        self.saveConfig = True
        pg.QtGui.QApplication.processEvents()

    def drawfigMain(self):
        """ Draws the main amplitude and spectrogram plots and any segments on them.
        Has to do some work to get the axis labels correct.
        """
        self.amplPlot.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=self.datalength,endpoint=True),self.audiodata)
        self.specPlot.setImage(self.sg)
        self.setColourMap(self.config['cmap'])
        self.setColourLevels()

        # Sort out the spectrogram frequency axis
        # The constants here are divided by 1000 to get kHz, and then remember the top is sampleRate/2
        FreqRange = (self.maxFreq-self.minFreq)/1000.
        self.specaxis.setTicks([[(0,self.minFreq/1000.),(np.shape(self.sg)[1]/4,self.minFreq/1000.+FreqRange/4.),(np.shape(self.sg)[1]/2,self.minFreq/1000.+FreqRange/2.),(3*np.shape(self.sg)[1]/4,self.minFreq/1000.+3*FreqRange/4.),(np.shape(self.sg)[1],self.minFreq/1000.+FreqRange)]])
        self.specaxis.setLabel('kHz')

        self.updateOverview()
        self.textpos = np.shape(self.sg)[1] + self.config['textoffset']

        # If there are segments, show them
        for count in range(len(self.segments)):
            self.addSegment(self.segments[count][0], self.segments[count][1],self.segments[count][2],self.segments[count][3],self.segments[count][4],False)

        # This is the moving bar for the playback
        if not hasattr(self,'bar'):
            self.bar = pg.InfiniteLine(angle=90, movable=True, pen={'color': 'c', 'width': 3})
        self.p_spec.addItem(self.bar, ignoreBounds=True)
        self.bar.sigPositionChangeFinished.connect(self.barMoved)

    def updateRegion_spec(self):
        """ This is the listener for when a segment box is changed in the spectrogram.
        It updates the position of the matching box, and also the text within it.
        """

        sender = self.sender()
        i = 0
        while self.listRectanglesa2[i] != sender and i<len(self.listRectanglesa2):
            i = i+1
        if i>len(self.listRectanglesa2):
            print "segment not found!"
        else:
            if type(sender) == self.ROItype:
                x1 = self.convertSpectoAmpl(sender.pos()[0])
                x2 = self.convertSpectoAmpl(sender.pos()[0]+sender.size()[0])
                self.segments[i][2] = sender.pos()[1]
                self.segments[i][3] = sender.pos()[1]+sender.size()[1]
                self.listLabels[i].setPos(sender.pos()[0], self.textpos)
            else:
                x1 = self.convertSpectoAmpl(sender.getRegion()[0])
                x2 = self.convertSpectoAmpl(sender.getRegion()[1])
                self.listLabels[i].setPos(sender.getRegion()[0], self.textpos)
            self.listRectanglesa1[i].setRegion([x1,x2])

            self.segments[i][0] = x1
            self.segments[i][1] = x2

    def updateRegion_ampl(self):
        """ This is the listener for when a segment box is changed in the waveform plot.
        It updates the position of the matching box, and also the text within it.
        """
        sender = self.sender()
        i = 0
        while self.listRectanglesa1[i] != sender and i<len(self.listRectanglesa1):
            i = i+1
        if i>len(self.listRectanglesa1):
            print "segment not found!"
        else:
            x1 = self.convertAmpltoSpec(sender.getRegion()[0])
            x2 = self.convertAmpltoSpec(sender.getRegion()[1])

            if type(self.listRectanglesa2[i]) == self.ROItype:
                y1 = self.listRectanglesa2[i].pos().y()
                y2 = self.listRectanglesa2[i].size().y()
                self.listRectanglesa2[i].setPos(pg.Point(x1,y1))
                self.listRectanglesa2[i].setSize(pg.Point(x2-x1,y2))
            else:
                self.listRectanglesa2[i].setRegion([x1,x2])
            self.listLabels[i].setPos(x1,self.textpos)

            self.segments[i][0] = sender.getRegion()[0]
            self.segments[i][1] = sender.getRegion()[1]

    def addSegment(self,startpoint,endpoint,y1=0,y2=0,species=None,saveSeg=True):
        """ When a new segment is created, does the work of creating it and connecting its
        listeners. Also updates the relevant overview segment.
        x, y are in amplitude coordinates.
        saveSeg means that we are drawing the saved ones. Need to check that those ones fit into
        the current window, can assume the other do, but have to save their times correctly."""


        if not saveSeg:
            timeRangeStart = self.startRead #self.currentFileSection*self.config['maxFileShow']
            timeRangeEnd = min(self.startRead + self.lenRead, float(self.fileLength) / self.sampleRate)
            print timeRangeStart, timeRangeEnd
            if startpoint > timeRangeStart and endpoint < timeRangeEnd:
                show = True
                # Put the startpoint and endpoint in the right range
                startpoint = startpoint - timeRangeStart
                endpoint = endpoint - timeRangeStart
            else:
                show = False
        else:
            show = True

        if show:
            # This is one we want to show


            # Get the name and colour sorted
            if species is None:
                species = "Don't Know"

            if species != "Don't Know":
                # Work out which overview segment this segment is in (could be more than one)
                inds = int(float(self.convertAmpltoSpec(startpoint))/self.widthOverviewSegment)
                inde = int(float(self.convertAmpltoSpec(endpoint))/self.widthOverviewSegment)
                if species[-1] == '?':
                    brush = self.ColourPossible
                    self.overviewSegments[inds:inde + 1, 2] += 1
                else:
                    brush = self.ColourNamed
                    self.overviewSegments[inds:inde + 1, 1] += 1
                self.prevBoxCol = brush

                for box in range(inds, inde + 1):
                    if self.overviewSegments[box,0] > 0:
                        self.SegmentRects[box].setBrush(self.ColourNone)
                    elif self.overviewSegments[box,2] > 0:
                        self.SegmentRects[box].setBrush(self.ColourPossible)
                    elif self.overviewSegments[box,1] > 0:
                        self.SegmentRects[box].setBrush(self.ColourNamed)
                    else:
                        self.SegmentRects[box].setBrush(pg.mkBrush('w'))
            else:
                brush = self.ColourNone
                self.prevBoxCol = brush
                # Work out which overview segment this segment is in (could be more than one)
                inds = int(float(self.convertAmpltoSpec(startpoint)) / self.widthOverviewSegment)
                inde = int(float(self.convertAmpltoSpec(endpoint)) / self.widthOverviewSegment)
                self.overviewSegments[inds:inde+1,0] += 1
                # Turn the colour of these segments in the overview
                for box in range(inds, inde + 1):
                    self.SegmentRects[box].setBrush(pg.mkBrush('w'))
                    self.SegmentRects[box].setBrush(self.ColourNone)
                    self.SegmentRects[box].update()

            # Make sure startpoint and endpoint are in the right order
            if startpoint > endpoint:
                temp = startpoint
                startpoint = endpoint
                endpoint = temp
            if y1 > y2:
                temp = y1
                y1 = y2
                y2 = temp

            # Add the segment in both plots and connect up the listeners
            p_ampl_r = pg.LinearRegionItem(brush=brush)
            self.p_ampl.addItem(p_ampl_r, ignoreBounds=True)
            p_ampl_r.setRegion([startpoint, endpoint])
            p_ampl_r.sigRegionChangeFinished.connect(self.updateRegion_ampl)

            if y1==0 and y2==0:
                p_spec_r = pg.LinearRegionItem(brush = brush)
                p_spec_r.setRegion([self.convertAmpltoSpec(startpoint), self.convertAmpltoSpec(endpoint)])
            else:
                startpointS = QPointF(self.convertAmpltoSpec(startpoint),y1)
                endpointS = QPointF(self.convertAmpltoSpec(endpoint),y2)
                p_spec_r = SupportClasses.ShadedRectROI(startpointS, endpointS - startpointS)
                if self.dragRectTransparent.isChecked():
                    col = self.prevBoxCol.rgb()
                    col = QtGui.QColor(col)
                    col.setAlpha(255)
                    p_spec_r.setBrush(None)
                    p_spec_r.setPen(pg.mkPen(col,width=1))
                else:
                    p_spec_r.setBrush(pg.mkBrush(self.prevBoxCol))
                    p_spec_r.setPen(pg.mkPen(None))
            self.p_spec.addItem(p_spec_r, ignoreBounds=True)
            p_spec_r.sigRegionChangeFinished.connect(self.updateRegion_spec)

            # Put the text into the box
            label = pg.TextItem(text=species, color='k')
            self.p_spec.addItem(label)
            label.setPos(self.convertAmpltoSpec(startpoint), self.textpos)

            # Add the segments to the relevent lists
            self.listRectanglesa1.append(p_ampl_r)
            self.listRectanglesa2.append(p_spec_r)
            self.listLabels.append(label)

            if saveSeg:
                # Add the segment to the data
                # Increment the time to be correct for the current section of the file
                self.segments.append([startpoint+self.startRead, endpoint+self.startRead, y1, y2, species])

    def mouseMoved(self,evt):
        """ Listener for mouse moves.
        If the user moves the mouse in the spectrogram, print the time, frequency, power for the mouse location. """
        if self.p_spec.sceneBoundingRect().contains(evt):
            mousePoint = self.p_spec.mapSceneToView(evt)
            indexx = int(mousePoint.x())
            indexy = int(mousePoint.y())
            if indexx > 0 and indexx < np.shape(self.sg)[0] and indexy > 0 and indexy < np.shape(self.sg)[1]:
                seconds = self.convertSpectoAmpl(mousePoint.x()) % 60
                minutes = int((self.convertSpectoAmpl(mousePoint.x()) / 60) % 60)
                self.pointData.setText('time=%d:%0.2f (m:s), freq=%0.1f (Hz),power=%0.1f (dB)' % (minutes,seconds, mousePoint.y() * self.sampleRate / 2. / np.shape(self.sg)[1], self.sg[indexx, indexy]))

    def mouseClicked_ampl(self,evt):
        """ Listener for if the user clicks on the amplitude plot.
        If there is a box selected, get its colour.
        If the user has clicked inside the scene, they could be
        (1) clicking in an already existing box -> select it
        (2) clicking anywhere else, or right-clicking in a box without having started a box -> start a box
        (3) clicking a second time to finish a box -> create the segment
        """
        pos = evt.scenePos()

        if self.box1id>-1:
            self.listRectanglesa1[self.box1id].setBrush(self.prevBoxCol)
            self.listRectanglesa1[self.box1id].update()
            if self.dragRectTransparent.isChecked() and type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                col = self.prevBoxCol.rgb()
                col = QtGui.QColor(col)
                col.setAlpha(255)
                #print "ampl", self.prevBoxCol.getRgb()
                self.listRectanglesa2[self.box1id].setPen(col,width=1)
            else:
                self.listRectanglesa2[self.box1id].setBrush(self.prevBoxCol)

            self.listRectanglesa2[self.box1id].update()

        if self.p_ampl.sceneBoundingRect().contains(pos):
            mousePoint = self.p_ampl.mapSceneToView(pos)

            if self.started:
                # This is the second click, so should pay attention and close the segment
                # Stop the mouse motion connection, remove the drawing boxes
                if self.started_window=='a':
                    self.p_ampl.scene().sigMouseMoved.disconnect()
                    self.p_ampl.removeItem(self.vLine_a)
                else:
                    self.p_spec.scene().sigMouseMoved.disconnect()
                    # Add the other mouse move listener back
                    self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
                    self.p_spec.removeItem(self.vLine_s)

                self.p_ampl.removeItem(self.drawingBox_ampl)
                self.p_spec.removeItem(self.drawingBox_spec)
                # If the user has pressed shift, copy the last species and don't use the context menu
                # If they pressed Control, add ? to the names
                modifiers = QtGui.QApplication.keyboardModifiers()
                if modifiers == QtCore.Qt.ShiftModifier:
                    self.addSegment(self.start_location, mousePoint.x(),species=self.lastSpecies)
                elif modifiers == QtCore.Qt.ControlModifier:
                    self.addSegment(self.start_location,mousePoint.x())
                    # Context menu
                    self.box1id = len(self.segments) - 1
                    self.fillBirdList(unsure=True)
                    self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))
                else:
                    self.addSegment(self.start_location,mousePoint.x())
                    # Context menu
                    self.box1id = len(self.segments) - 1
                    self.fillBirdList()
                    self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))

                self.playSegButton.setEnabled(True)
                self.playBandLimitedSegButton.setEnabled(True)

                self.listRectanglesa1[self.box1id].setBrush(fn.mkBrush(self.ColourSelected))
                self.listRectanglesa1[self.box1id].update()

                if self.dragRectTransparent.isChecked() and type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                    self.listRectanglesa2[self.box1id].setPen(fn.mkPen(self.ColourSelectedDark,width=1))
                else:
                    self.listRectanglesa2[self.box1id].setBrush(fn.mkBrush(self.ColourSelected))

                self.listRectanglesa2[self.box1id].update()

                self.started = not(self.started)
            else:
                # Check if the user has clicked in a box
                # Note: Returns the first one it finds
                box1id = -1
                for count in range(len(self.listRectanglesa1)):
                    x1, x2 = self.listRectanglesa1[count].getRegion()
                    if x1 <= mousePoint.x() and x2 >= mousePoint.x():
                        box1id = count

                if box1id > -1 and not evt.button() == QtCore.Qt.RightButton:
                    # User clicked in a box (with the left button)
                    # Change colour, store the old colour
                    self.box1id = box1id
                    self.prevBoxCol = self.listRectanglesa1[box1id].brush.color()
                    self.listRectanglesa1[box1id].setBrush(fn.mkBrush(self.ColourSelected))
                    self.listRectanglesa1[box1id].update()
                    self.playSegButton.setEnabled(True)
                    self.playBandLimitedSegButton.setEnabled(True)
                    if self.dragRectTransparent.isChecked() and type(self.listRectanglesa2[box1id]) == self.ROItype:
                        self.listRectanglesa2[box1id].setPen(fn.mkPen(self.ColourSelectedDark,width=1))
                    else:
                        self.listRectanglesa2[box1id].setBrush(fn.mkBrush(self.ColourSelected))

                    self.listRectanglesa2[box1id].update()

                    modifiers = QtGui.QApplication.keyboardModifiers()
                    if modifiers == QtCore.Qt.ControlModifier:
                        self.fillBirdList(unsure=True)
                    else:
                        self.fillBirdList()
                    self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))
                else:
                    # User hasn't clicked in a box (or used the right button), so start a new segment
                    self.start_location = mousePoint.x()
                    self.vLine_a = pg.InfiniteLine(angle=90, movable=False,pen={'color': 'r', 'width': 3})
                    self.p_ampl.addItem(self.vLine_a, ignoreBounds=True)
                    self.vLine_a.setPos(self.start_location)

                    self.playSegButton.setEnabled(False)
                    self.playBandLimitedSegButton.setEnabled(True)
                    brush = self.ColourNone
                    self.drawingBox_ampl = pg.LinearRegionItem(brush=brush)
                    self.p_ampl.addItem(self.drawingBox_ampl, ignoreBounds=True)
                    self.drawingBox_ampl.setRegion([self.start_location, self.start_location])
                    self.drawingBox_spec = pg.LinearRegionItem(brush=brush)
                    self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                    self.drawingBox_spec.setRegion([self.convertAmpltoSpec(self.start_location), self.convertAmpltoSpec(self.start_location)])
                    self.p_ampl.scene().sigMouseMoved.connect(self.GrowBox_ampl)
                    self.started_window = 'a'

                    self.started = not (self.started)

    def mouseClicked_spec(self,evt):
        """ Listener for if the user clicks on the spectrogram plot.
        See the amplitude version (mouseClicked_ampl()) for details. Although much of the code is a repeat,
        it is separated for clarity.
        """
        pos = evt.scenePos()

        if self.box1id>-1:
            self.listRectanglesa1[self.box1id].setBrush(self.prevBoxCol)
            self.listRectanglesa1[self.box1id].update()
            if self.dragRectTransparent.isChecked() and type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                col = self.prevBoxCol.rgb()
                col = QtGui.QColor(col)
                col.setAlpha(255)
                self.listRectanglesa2[self.box1id].setPen(col,width=1)
            else:
                self.listRectanglesa2[self.box1id].setBrush(self.prevBoxCol)
            self.listRectanglesa2[self.box1id].update()

        if self.p_spec.sceneBoundingRect().contains(pos):
            mousePoint = self.p_spec.mapSceneToView(pos)

            if self.started:
                # This is the second click, so should pay attention and close the segment
                # Stop the mouse motion connection, remove the drawing boxes
                if self.dragRectangles.isChecked():
                    return
                else:
                    if self.started_window == 's':
                        self.p_spec.scene().sigMouseMoved.disconnect()
                        self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
                        self.p_spec.removeItem(self.vLine_s)
                    else:
                        self.p_ampl.scene().sigMouseMoved.disconnect()
                        self.p_ampl.removeItem(self.vLine_a)
                    self.p_ampl.removeItem(self.drawingBox_ampl)
                    self.p_spec.removeItem(self.drawingBox_spec)
                    # If the user has pressed shift, copy the last species and don't use the context menu
                    modifiers = QtGui.QApplication.keyboardModifiers()
                    if modifiers == QtCore.Qt.ShiftModifier:
                        self.addSegment(self.start_location, self.convertSpectoAmpl(mousePoint.x()), species=self.lastSpecies)
                    elif modifiers == QtCore.Qt.ControlModifier:
                        self.addSegment(self.start_location, self.convertSpectoAmpl(mousePoint.x()))
                        # Context menu
                        self.box1id = len(self.segments) - 1
                        self.fillBirdList(unsure=True)
                        self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))
                    else:
                        self.addSegment(self.start_location, self.convertSpectoAmpl(mousePoint.x()))
                        # Context menu
                        self.box1id = len(self.segments) - 1
                        self.fillBirdList()
                        self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))

                    self.playSegButton.setEnabled(True)
                    self.playBandLimitedSegButton.setEnabled(True)

                    self.listRectanglesa1[self.box1id].setBrush(fn.mkBrush(self.ColourSelected))
                    self.listRectanglesa1[self.box1id].update()
                    if self.dragRectTransparent.isChecked() and type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                        self.listRectanglesa2[self.box1id].setPen(fn.mkPen(self.ColourSelectedDark,width=1))
                    else:
                        self.listRectanglesa2[self.box1id].setBrush(fn.mkBrush(self.ColourSelected))

                    self.listRectanglesa2[self.box1id].update()

                    self.started = not(self.started)
            else:
                # Check if the user has clicked in a box
                # Note: Returns the first one it finds
                box1id = -1
                for count in range(len(self.listRectanglesa2)):
                    if type(self.listRectanglesa2[count]) == self.ROItype:
                        x1 = self.listRectanglesa2[count].pos().x()
                        y1 = self.listRectanglesa2[count].pos().y()
                        x2 = x1 + self.listRectanglesa2[count].size().x()
                        y2 = y1 + self.listRectanglesa2[count].size().y()
                        if x1 <= mousePoint.x() and x2 >= mousePoint.x() and y1 <= mousePoint.y() and y2 >= mousePoint.y():
                            box1id = count
                    else:
                        x1, x2 = self.listRectanglesa2[count].getRegion()
                        if x1 <= mousePoint.x() and x2 >= mousePoint.x():
                            box1id = count

                if box1id > -1 and not evt.button() == QtCore.Qt.RightButton:
                    # User clicked in a box (with the left button)
                    self.box1id = box1id
                    self.prevBoxCol = self.listRectanglesa1[box1id].brush.color()
                    self.listRectanglesa1[box1id].setBrush(fn.mkBrush(self.ColourSelected))
                    self.listRectanglesa1[box1id].update()
                    self.playSegButton.setEnabled(True)
                    self.playBandLimitedSegButton.setEnabled(True)
                    if self.dragRectTransparent.isChecked() and type(self.listRectanglesa2[box1id]) == self.ROItype:
                        self.listRectanglesa2[box1id].setPen(fn.mkPen(self.ColourSelectedDark,width=1))
                    else:
                        self.listRectanglesa2[box1id].setBrush(fn.mkBrush(self.ColourSelected))

                    self.listRectanglesa2[box1id].update()
                    modifiers = QtGui.QApplication.keyboardModifiers()
                    if modifiers == QtCore.Qt.ControlModifier:
                        self.fillBirdList(unsure=True)
                    else:
                        self.fillBirdList()
                    self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))
                else:
                    # User hasn't clicked in a box (or used the right button), so start a new segment
                    # Note that need to click in the same plot both times.
                    if self.dragRectangles.isChecked():
                        return
                    else:
                        self.start_location = self.convertSpectoAmpl(mousePoint.x())
                        self.vLine_s = pg.InfiniteLine(angle=90, movable=False,pen={'color': 'r', 'width': 3})
                        self.p_spec.addItem(self.vLine_s, ignoreBounds=True)
                        self.vLine_s.setPos(mousePoint.x())
                        self.playSegButton.setEnabled(False)
                        self.playBandLimitedSegButton.setEnabled(True)

                        brush = self.ColourNone
                        self.drawingBox_ampl = pg.LinearRegionItem(brush=brush)
                        self.p_ampl.addItem(self.drawingBox_ampl, ignoreBounds=True)
                        self.drawingBox_ampl.setRegion([self.start_location, self.start_location])
                        self.drawingBox_spec = pg.LinearRegionItem(brush=brush)
                        self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                        self.drawingBox_spec.setRegion([mousePoint.x(),mousePoint.x()])
                        self.p_spec.scene().sigMouseMoved.connect(self.GrowBox_spec)
                        self.started_window = 's'

                        self.started = not (self.started)

    def mouseDragged_spec(self, evt1, evt2, evt3):
        """ Listener for if the user drags in the spectrogram plot.
        It's a bit simpler than the click ones, since there is less ambiguity.
        Again, some of the code is a repeat, but kept self-contained for ease. """
        print "in mouse dragged"
        if self.box1id>-1:
            self.listRectanglesa1[self.box1id].setBrush(self.prevBoxCol)
            self.listRectanglesa1[self.box1id].update()
            if self.dragRectTransparent.isChecked() and type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                col = self.prevBoxCol.rgb()
                col = QtGui.QColor(col)
                col.setAlpha(255)
                self.listRectanglesa2[self.box1id].setPen(pg.mkPen(col,width=1))
            else:
                self.listRectanglesa2[self.box1id].setBrush(self.prevBoxCol)
            self.listRectanglesa2[self.box1id].update()

        if self.dragRectangles.isChecked():
            evt1 = self.p_spec.mapSceneToView(evt1)
            evt2 = self.p_spec.mapSceneToView(evt2)

            # If the user has pressed shift, copy the last species and don't use the context menu
            modifiers = QtGui.QApplication.keyboardModifiers()
            if modifiers == QtCore.Qt.ShiftModifier:
                self.addSegment(self.convertSpectoAmpl(evt1.x()), self.convertSpectoAmpl(evt2.x()), evt1.y(), evt2.y(),self.lastSpecies)
            elif modifiers == QtCore.Qt.ControlModifier:
                self.addSegment(self.convertSpectoAmpl(evt1.x()), self.convertSpectoAmpl(evt2.x()), evt1.y(), evt2.y())
                # Context menu
                self.box1id = len(self.segments) - 1
                self.fillBirdList(unsure=True)
                self.menuBirdList.popup(QPoint(evt3.x(), evt3.y()))
            else:
                self.addSegment(self.convertSpectoAmpl(evt1.x()), self.convertSpectoAmpl(evt2.x()), evt1.y(), evt2.y())
                # Context menu
                self.box1id = len(self.segments) - 1
                self.fillBirdList()
                self.menuBirdList.popup(QPoint(evt3.x(), evt3.y()))

            self.playSegButton.setEnabled(True)
            self.playBandLimitedSegButton.setEnabled(True)

            self.listRectanglesa1[self.box1id].setBrush(fn.mkBrush(self.ColourSelected))
            self.listRectanglesa1[self.box1id].update()
            if self.dragRectTransparent.isChecked() and type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                self.listRectanglesa2[self.box1id].setBrush(fn.mkBrush(None))
                self.listRectanglesa2[self.box1id].setPen(fn.mkPen(self.ColourSelectedDark,width=1))
            else:
                self.listRectanglesa2[self.box1id].setBrush(fn.mkBrush(self.ColourSelected))
                self.listRectanglesa2[self.box1id].setPen(None)

            self.listRectanglesa2[self.box1id].update()
        else:
            return

    def GrowBox_ampl(self,evt):
        """ Listener for when a segment is being made in the amplitude plot.
        Makes the blue box that follows the mouse change size. """
        pos = evt
        if self.p_ampl.sceneBoundingRect().contains(pos):
            mousePoint = self.p_ampl.mapSceneToView(pos)
            self.drawingBox_ampl.setRegion([self.start_location, mousePoint.x()])
            self.drawingBox_spec.setRegion([self.convertAmpltoSpec(self.start_location), self.convertAmpltoSpec(mousePoint.x())])

    def GrowBox_spec(self, evt):
        """ Listener for when a segment is being made in the spectrogram plot.
        Makes the blue box that follows the mouse change size. """
        pos = evt
        if self.p_spec.sceneBoundingRect().contains(pos):
            mousePoint = self.p_spec.mapSceneToView(pos)
            self.drawingBox_ampl.setRegion([self.start_location, self.convertSpectoAmpl(mousePoint.x())])
            self.drawingBox_spec.setRegion([self.convertAmpltoSpec(self.start_location), mousePoint.x()])

    def birdSelected(self,birdname,update=True):
        """ Collects the label for a bird from the context menu and processes it.
        Has to update the overview segments in case their colour should change.
        Also handles getting the name through a message box if necessary.
        """
        oldname = self.segments[self.box1id][4]
        # Work out which overview segment this segment is in (could be more than one)
        inds = int(float(self.convertAmpltoSpec(self.segments[self.box1id][0])) / self.widthOverviewSegment)
        inde = int(float(self.convertAmpltoSpec(self.segments[self.box1id][1])) / self.widthOverviewSegment)
        if oldname == "Don't Know":
            if birdname != "Don't Know":
                if birdname[-1] == '?':
                    self.overviewSegments[inds:inde + 1, 0] -= 1
                    self.overviewSegments[inds:inde + 1, 2] += 1
                else:
                    self.overviewSegments[inds:inde + 1, 0] -= 1
                    self.overviewSegments[inds:inde + 1, 1] += 1
        elif oldname[-1] == '?':
            if birdname[-1] != '?':
                if birdname == "Don't Know":
                    self.overviewSegments[inds:inde + 1, 2] -= 1
                    self.overviewSegments[inds:inde + 1, 0] += 1
                else:
                    self.overviewSegments[inds:inde + 1, 2] -= 1
                    self.overviewSegments[inds:inde + 1, 1] += 1
        else:
            if birdname == "Don't Know":
                self.overviewSegments[inds:inde + 1, 1] -= 1
                self.overviewSegments[inds:inde + 1, 0] += 1
            elif birdname[-1] == '?':
                self.overviewSegments[inds:inde + 1, 1] -= 1
                self.overviewSegments[inds:inde + 1, 2] += 1

        for box in range(inds, inde + 1):
            if self.overviewSegments[box, 0] > 0:
                self.SegmentRects[box].setBrush(pg.mkBrush('w'))
                self.SegmentRects[box].setBrush(self.ColourNone)
                self.SegmentRects[box].update()
            elif self.overviewSegments[box, 2] > 0:
                self.SegmentRects[box].setBrush(self.ColourPossible)
            elif self.overviewSegments[box, 1] > 0:
                self.SegmentRects[box].setBrush(pg.mkBrush('w'))
                self.SegmentRects[box].setBrush(self.ColourNamed)
                self.SegmentRects[box].update()
            else:
                self.SegmentRects[box].setBrush(pg.mkBrush('w'))

        # Now update the text
        if birdname is not 'Other':
            self.updateText(birdname)
            if update:
                # Put the selected bird name at the top of the list
                if birdname[-1] == '?':
                    birdname = birdname[:-1]
                self.config['BirdList'].remove(birdname)
                self.config['BirdList'].insert(0,birdname)
        else:
            text, ok = QInputDialog.getText(self, 'Bird name', 'Enter the bird name:')
            if ok:
                text = str(text).title()
                self.updateText(text)

                if text in self.config['BirdList']:
                    pass
                else:
                    # Add the new bird name.
                    if update:
                        self.config['BirdList'].insert(0,text)
                    else:
                        self.config['BirdList'].append(text)
                    self.saveConfig = True

    def updateText(self, text,segID=None):
        """ When the user sets or changes the name in a segment, update the text and the colour. """
        if segID is None:
            segID = self.box1id
        self.segments[segID][4] = text
        self.listLabels[segID].setText(text,'k')

        # Update the colour
        if text != "Don't Know":
            if text[-1] == '?':
                self.prevBoxCol = self.ColourPossible
            else:
                self.prevBoxCol = self.ColourNamed
        else:
            self.prevBoxCol = self.ColourNone

        # Store the species in case the user wants it for the next segment
        self.lastSpecies = text

    def setColourMap(self,cmap):
        """ Listener for the menu item that chooses a colour map.
        Loads them from the file as appropriate and sets the lookup table.
        """
        self.config['cmap'] = cmap

        import colourMaps
        pos, colour, mode = colourMaps.colourMaps(cmap)

        cmap = pg.ColorMap(pos, colour,mode)
        self.lut = cmap.getLookupTable(0.0, 1.0, 256)

        self.specPlot.setLookupTable(self.lut)
        self.overviewImage.setLookupTable(self.lut)

    def invertColourMap(self):
        """ Listener for the menu item that converts the colour map"""
        self.config['invertColourMap'] = not self.config['invertColourMap']
        self.setColourLevels()

    def setColourLevels(self):
        """ Listener for the brightness and contrast sliders being changed. Also called when spectrograms are loaded, etc.
        Translates the brightness and contrast values into appropriate image levels. """
        minsg = np.min(self.sg)
        maxsg = np.max(self.sg)
        self.config['brightness'] = self.brightnessSlider.value()
        self.config['contrast'] = self.contrastSlider.value()
        self.colourStart = (self.config['brightness'] / 100.0 * self.config['contrast'] / 100.0) * (maxsg - minsg) + minsg
        self.colourEnd = (maxsg - minsg) * (1.0 - self.config['contrast'] / 100.0) + self.colourStart

        if self.config['invertColourMap']:
            self.overviewImage.setLevels([self.colourEnd, self.colourStart])
            self.specPlot.setLevels([self.colourEnd, self.colourStart])
        else:
            self.overviewImage.setLevels([self.colourStart, self.colourEnd])
            self.specPlot.setLevels([self.colourStart, self.colourEnd])

    def moveLeft(self):
        """ When the left button is pressed (next to the overview plot), move everything along
        Allows a 10% overlap """
        minX, maxX = self.overviewImageRegion.getRegion()
        newminX = max(0,minX-(maxX-minX)*0.9)
        self.overviewImageRegion.setRegion([newminX, newminX+maxX-minX])
        self.updateOverview()
        self.playPosition = int(self.convertSpectoAmpl(newminX)*1000.0)

    def moveRight(self):
        """ When the right button is pressed (next to the overview plot), move everything along
        Allows a 10% overlap """
        minX, maxX = self.overviewImageRegion.getRegion()
        newminX = min(np.shape(self.sg)[0]-(maxX-minX),minX+(maxX-minX)*0.9)
        self.overviewImageRegion.setRegion([newminX, newminX+maxX-minX])
        self.updateOverview()
        self.playPosition = int(self.convertSpectoAmpl(newminX)*1000.0)

    def prepare5minMove(self):
        if self.segments != [] or self.hasSegments:
            if len(self.segments)>0:
                if self.segments[0][0] > -1:
                    self.segments.insert(0, [-1, -1, str(self.username), -1, -1])
            else:
                self.segments.insert(0, [-1, -1, str(self.username), -1, -1])
            self.saveSegments()
        self.resetStorageArrays()
        # Reset the media player
        if self.media_obj.state() == phonon.Phonon.PlayingState:
            self.media_obj.pause()
            self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))
        self.loadFile()


    def movePrev5mins(self):
        """ When the button to move to the next 5 minutes is pressed, enable that.
        Have to update the timeaxis offset, check if the buttons should be disabled or not,
        save the segments and reset the arrays, then call loadFile.
        """
        self.currentFileSection -= 1
        self.next5mins.setEnabled(True)
        if self.currentFileSection == 0:
            self.prev5mins.setEnabled(False)
        self.prepare5minMove()


    def moveNext5mins(self):
        """ When the button to move to the previous 5 minutes is pressed, enable that.
        Have to update the timeaxis offset, check if the buttons should be disabled or not,
        save the segments and reset the arrays, then call loadFile.
        """
        self.currentFileSection += 1
        self.prev5mins.setEnabled(True)
        if self.currentFileSection == self.nFileSections-1:
            self.next5mins.setEnabled(False)
        self.prepare5minMove()

    def scroll(self):
        """ When the slider at the bottom of the screen is moved, move everything along. """
        newminX = self.scrollSlider.value()
        minX, maxX = self.overviewImageRegion.getRegion()
        self.overviewImageRegion.setRegion([newminX, newminX+maxX-minX])
        self.updateOverview()
        self.playPosition = int(self.convertSpectoAmpl(newminX)*1000.0)

    def changeWidth(self, value):
        """ Listener for the spinbox that decides the width of the main window.
        It updates the top figure plots as the window width is changed.
        Slightly annoyingly, it gets called when the value gets reset, hence the first line. """
        # **** Needs work for paging

        if not hasattr(self,'overviewImageRegion'):
            return
        self.windowSize = value

        # Redraw the highlight in the overview figure appropriately
        minX, maxX = self.overviewImageRegion.getRegion()
        newmaxX = self.convertAmpltoSpec(value)+minX
        self.overviewImageRegion.setRegion([minX, newmaxX])
        self.scrollSlider.setMaximum(np.shape(self.sg)[0]-self.convertAmpltoSpec(self.widthWindow.value()))
        self.updateOverview()

# ===============
# Generate the various dialogs that match the menu items

    def humanClassifyDialog1(self):
        # Create the dialog that shows calls to the user for verification
        if len(self.segments)==0:
            print "No box selected"
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("No segments to check")
            msg.setWindowIcon(QIcon('img/Avianz.ico'))
            msg.setWindowTitle("No segments")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            return
        else:
            # self.statusLeft.setText("Checking...")
            self.box1id = 0
            # Different calls for the two types of region
            if type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                x1 = self.listRectanglesa2[self.box1id].pos()[0]
                x2 = x1 + self.listRectanglesa2[self.box1id].size()[0]
            else:
                x1, x2 = self.listRectanglesa2[self.box1id].getRegion()
            x1 = int(x1)
            x2 = int(x2)
            x3 = int(self.listRectanglesa1[self.box1id].getRegion()[0] * self.sampleRate)
            x4 = int(self.listRectanglesa1[self.box1id].getRegion()[1] * self.sampleRate)
            self.humanClassifyDialog1 = Dialogs.HumanClassify1(self.sg[x1:x2,:],self.audiodata[x3:x4],self.sampleRate,self.segments[self.box1id][4],self.lut,self.colourStart,self.colourEnd,self.config['invertColourMap'], self.config['BirdList'])
            self.humanClassifyDialog1.show()
            self.humanClassifyDialog1.activateWindow()
            #self.humanClassifyDialog1.close.clicked.connect(self.humanClassifyClose1)
            self.humanClassifyDialog1.correct.clicked.connect(self.humanClassifyCorrect1)
            self.humanClassifyDialog1.delete.clicked.connect(self.humanClassifyDelete1)
            # self.statusLeft.setText("Ready")

    def humanClassifyClose1(self):
        # Listener for the human verification dialog.
        self.humanClassifyDialog1.done(1)

    def humanClassifyNextImage1(self):
        # Get the next image
        if self.box1id != len(self.listRectanglesa2)-1:
            self.box1id += 1
            # Different calls for the two types of region
            if type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                x1 = self.listRectanglesa2[self.box1id].pos()[0]
                x2 = x1 + self.listRectanglesa2[self.box1id].size()[0]
            else:
                x1, x2 = self.listRectanglesa2[self.box1id].getRegion()
            x1 = int(x1)
            x2 = int(x2)
            x3 = int(self.listRectanglesa1[self.box1id].getRegion()[0] * self.sampleRate)
            x4 = int(self.listRectanglesa1[self.box1id].getRegion()[1] * self.sampleRate)
            self.humanClassifyDialog1.setImage(self.sg[x1:x2,:],self.audiodata[x3:x4],self.sampleRate,self.segments[self.box1id][4])
        else:
            print "Last image"
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowIcon(QIcon('Avianz.ico'))
            msg.setText("All segmentations checked")
            msg.setWindowTitle("Finished")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            self.humanClassifyClose1()

    def humanClassifyCorrect1(self):
        label, self.saveConfig, checkText = self.humanClassifyDialog1.getValues()
        if len(checkText) > 0:
            if label != checkText:
                label = str(checkText)
                self.humanClassifyDialog1.birdTextEntered()
                self.saveConfig = True
            #self.humanClassifyDialog1.tbox.setText('')

        if label != self.segments[self.box1id][4]:
            #self.updateText(label)
            self.birdSelected(label,update=False)

            self.listRectanglesa1[self.box1id].setBrush(self.prevBoxCol)
            self.listRectanglesa1[self.box1id].update()
            if self.dragRectTransparent.isChecked() and type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                col = self.prevBoxCol.rgb()
                col = QtGui.QColor(col)
                col.setAlpha(255)
                #print "ampl", self.prevBoxCol.getRgb()
                self.listRectanglesa2[self.box1id].setPen(col,width=1)
            else:
                self.listRectanglesa2[self.box1id].setBrush(self.prevBoxCol)

            self.listRectanglesa2[self.box1id].update()

            if self.saveConfig:
                self.config['BirdList'].append(label)
        self.humanClassifyDialog1.tbox.setText('')
        self.humanClassifyDialog1.tbox.setEnabled(False)
        self.humanClassifyNextImage1()

    def humanClassifyDelete1(self):
        # Delete a segment
        id = self.box1id
        self.deleteSegment(self.box1id)
        self.box1id = id-1
        self.humanClassifyNextImage1()

    def humanClassifyDialog2(self):
        # Create the dialog that shows calls to the user for verification
        if len(self.segments)==0:
            print "No box selected"
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("No segments to check")
            msg.setWindowIcon(QIcon('img/Avianz.ico'))
            msg.setWindowTitle("No segment")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            return
        self.statusLeft.setText("Checking...")
        if len(self.segments)>0:
            names = [item[4] for item in self.segments]
            names = [n if n[-1] != '?' else n[:-1] for n in names]
            # Should make them unique
            keys = {}
            for n in names:
                keys[n] = 1
            names = keys.keys()
            self.humanClassifyDialog2a = Dialogs.HumanClassify2a(names)

            if self.humanClassifyDialog2a.exec_() == 1:
                label = self.humanClassifyDialog2a.getValues()
                self.humanClassifyDialog2 = Dialogs.HumanClassify2(self.sg,self.segments,label,self.sampleRate, self.config['incr'], self.lut,self.colourStart,self.colourEnd,self.config['invertColourMap'])
                #self.humanClassifyDialog2.show()
                self.humanClassifyDialog2.exec_()
                errors = self.humanClassifyDialog2.getValues()
                #print errors
                # TODO: Should store these somewhere and improve the learning, for now just deleting
                for error in errors[-1::-1]:
                    self.deleteSegment(error)
        self.statusLeft.setText("Ready")

    def showSpectrogramDialog(self):
        # Create the spectrogram dialog when the button is pressed
        if not hasattr(self,'spectrogramDialog'):
            self.spectrogramDialog = Dialogs.Spectrogram(self.config['window_width'],self.config['incr'])
        self.spectrogramDialog.show()
        self.spectrogramDialog.activateWindow()
        self.spectrogramDialog.activate.clicked.connect(self.spectrogram)
        # TODO: next line
        # self.connect(self.spectrogramDialog, SIGNAL("changed"), self.spectrogram)

    def spectrogram(self):
        # Listener for the spectrogram dialog.
        [alg, mean_normalise, multitaper, window_width, incr] = self.spectrogramDialog.getValues()
        self.statusLeft.setText("Updating the spectrogram...")
        self.sp.set_width(int(str(window_width)), int(str(incr)))
        sgRaw = self.sp.spectrogram(self.audiodata,str(alg),mean_normalise=True,onesided=True,multitaper=multitaper)
        maxsg = np.min(sgRaw)
        self.sg = np.abs(np.where(sgRaw==0,0.0,10.0 * np.log10(sgRaw/maxsg)))

        #self.amplPlot.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=self.datalength,endpoint=True),self.audiodata)
        #self.w_spec.removeItem(self.timeaxis)
        #self.timeaxis = None

        # If the size of the spectrogram has changed, need to update the positions of things
        if int(str(incr)) != self.config['incr'] or int(str(window_width)) != self.config['window_width']:
            self.config['incr'] = int(str(incr))
            self.config['window_width'] = int(str(window_width))
            self.changeWidth(self.widthWindow.value())
            # Update the positions of the segments
            self.textpos = np.shape(self.sg)[1] + self.config['textoffset']
            for s in range(len(self.listRectanglesa2)):
                x1 = self.convertAmpltoSpec(self.listRectanglesa1[s].getRegion()[0])
                x2 = self.convertAmpltoSpec(self.listRectanglesa1[s].getRegion()[1])
                self.listRectanglesa2[s].setRegion([x1, x2])
                self.listLabels[s].setPos(x1,self.textpos)
            # Update the axis
            FreqRange = (self.maxFreq - self.minFreq)/1000.
            self.specaxis.setTicks([[(0, self.minFreq / 1000.),
                                     (np.shape(self.sg)[1] / 4, self.minFreq / 1000. + FreqRange / 4.),
                                     (np.shape(self.sg)[1] / 2, self.minFreq / 1000. + FreqRange / 2.),
                                     (3 * np.shape(self.sg)[1] / 4, self.minFreq / 1000. + 3 * FreqRange / 4.),
                                     (np.shape(self.sg)[1], self.minFreq / 1000. + FreqRange)]])

            # Redraw everything and redraw it
            self.removeSegments()
            for r in self.SegmentRects:
                self.p_overview2.removeItem(r)
            self.SegmentRects = []
            self.p_overview.removeItem(self.overviewImageRegion)

            self.drawOverview()
            self.drawfigMain()
            if hasattr(self, 'seg'):
                self.seg.setNewData(self.audiodata, sgRaw, self.sampleRate, self.config['window_width'],
                                    self.config['incr'])
        self.statusLeft.setText("Ready")

    def denoiseDialog(self):
        # Create the denoising dialog when the relevant button is pressed
        # TODO: Anything to help setting bandpass levels?
        self.denoiseDialog = Dialogs.Denoise()
        self.denoiseDialog.show()
        self.denoiseDialog.activateWindow()
        self.denoiseDialog.activate.clicked.connect(self.denoise)
        self.denoiseDialog.undo.clicked.connect(self.denoise_undo)
        self.denoiseDialog.save.clicked.connect(self.denoise_save)

    def backup(self):
        if hasattr(self, 'audiodata_backup'):
            if self.audiodata_backup is not None:
                audiodata_backup_new = np.empty(
                    (np.shape(self.audiodata_backup)[0], np.shape(self.audiodata_backup)[1] + 1))
                audiodata_backup_new[:, :-1] = np.copy(self.audiodata_backup)
                audiodata_backup_new[:, -1] = np.copy(self.audiodata)
                self.audiodata_backup = audiodata_backup_new
            else:
                self.audiodata_backup = np.empty((np.shape(self.audiodata)[0], 1))
                self.audiodata_backup[:, 0] = np.copy(self.audiodata)
        else:
            self.audiodata_backup = np.empty((np.shape(self.audiodata)[0], 1))
            self.audiodata_backup[:, 0] = np.copy(self.audiodata)

    def denoise(self):
        # Listener for the denoising dialog.
        # Calls the denoiser and then plots the updated data
        # TODO: should it be saved automatically, or a button added?
        [alg,depthchoice,depth,thrType,thr,wavelet,start,end,width] = self.denoiseDialog.getValues()
        # TODO: deal with these!
        # TODO: Undo needs testing
        self.backup()
        self.statusLeft.setText("Denoising...")
        if str(alg) == "Wavelets":
            if thrType is True:
                type = 'Soft'
            else:
                type = 'Hard'
            if depthchoice:
                depth = None
            else:
                depth = int(str(depth))
            self.audiodata = self.sp.waveletDenoise(self.audiodata,type,float(str(thr)),depth,wavelet=str(wavelet))
        elif str(alg) == "Bandpass --> Wavelets":
            if thrType is True:
                type = 'soft'
            else:
                type = 'hard'
            if depthchoice:
                depth = None
            else:
                depth = int(str(depth))
            self.audiodata = self.sp.bandpassFilter(self.audiodata,int(str(start)),int(str(end)))
            self.audiodata = self.sp.waveletDenoise(self.audiodata,type,float(str(thr)),depth,str(wavelet))
        elif str(alg) == "Wavelets --> Bandpass":
            if thrType is True:
                type = 'soft'
            else:
                type = 'hard'
            if depthchoice:
                depth = None
            else:
                depth = int(str(depth))
            self.audiodata = self.sp.waveletDenoise(self.audiodata,type,float(str(thr)),depth,str(wavelet))
            self.audiodata = self.sp.bandpassFilter(self.audiodata,int(str(start)),int(str(end)))

        elif str(alg) == "Bandpass":
            # self.audiodata = self.sp.bandpassFilter(self.audiodata, int(str(start)), int(str(end)))
            self.audiodata = self.sp.ButterworthBandpass(self.audiodata, self.sampleRate, low=int(str(start)), high=int(str(end)))
        elif str(alg) == "Butterworth Bandpass":
            self.audiodata = self.sp.ButterworthBandpass(self.audiodata, self.sampleRate, low=int(str(start)), high=int(str(end)))
        else:
            #"Median Filter"
            self.audiodata = self.sp.medianFilter(self.audiodata,int(str(width)))

        if platform.system() == 'Darwin':
            filename = 'temp.wav'
        else:
            import tempfile
            f = tempfile.NamedTemporaryFile(mode='w+t', delete=False)
            filename = f.name
        # print filename
        # self.audiodata = self.audiodata.astype('int16')
        wavio.write(filename,self.audiodata.astype('int16'),self.sampleRate,scale='dtype-limits',sampwidth=2)

        #open the temp file
        wavobj = wavio.read(filename)
        self.sampleRate = wavobj.rate
        self.audiodata = wavobj.data
        if self.audiodata.dtype is not 'float':
            self.audiodata = self.audiodata.astype('float') #/ 32768.0

        self.audiodata=self.audiodata[:,0]

        #f.close()

        # print np.shape(self.audiodata)
        sgRaw = self.sp.spectrogram(self.audiodata,self.sampleRate,mean_normalise=True,onesided=True,multitaper=False)
        maxsg = np.min(sgRaw)
        self.sg = np.abs(np.where(sgRaw==0,0.0,10.0 * np.log10(sgRaw/maxsg)))
        self.overviewImage.setImage(self.sg)

        self.specPlot.setImage(self.sg)
        self.amplPlot.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=self.datalength,endpoint=True),self.audiodata)
        # self.amplPlot.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate*1000.0,num=self.datalength,endpoint=True),self.audiodata)
        #self.specaxis.setTicks([[(0,0),(np.shape(self.sg)[1]/4,self.sampleRate/8000),(np.shape(self.sg)[1]/2,self.sampleRate/4000),(3*np.shape(self.sg)[1]/4,3*self.sampleRate/8000),(np.shape(self.sg)[1],self.sampleRate/2000)]])
        self.minFreq = int(str(start))
        self.maxFreq = int(str(end))

        self.setColourLevels()

        self.statusLeft.setText("Ready")

        # # media_obj = phonon.Phonon.MediaObject(self)
        # self.media_obj.setCurrentSource(phonon.Phonon.MediaSource(filename))
        # audio_output = phonon.Phonon.AudioOutput(phonon.Phonon.MusicCategory, self)
        # phonon.Phonon.createPath(self.media_obj, audio_output)
        # self.media_obj.tick.connect(self.movePlaySlider)
        # self.media_obj.setTickInterval(20)
        # self.media_obj.tick.connect(self.movePlaySlider)
        # self.media_obj.finished.connect(self.playFinished)

    def redoFreqAxis(self):
        start = self.minFreq
        end = self.maxFreq

        height = self.sampleRate / 2. / np.shape(self.sg)[1]

        self.overviewImage.setImage(self.sg[:,int(float(start)/height):int(float(end)/height)])
        self.specPlot.setImage(self.sg[:,int(float(start)/height):int(float(end)/height)])

        FreqRange = end - start
        SpecRange = FreqRange/height
        self.specaxis.setTicks([[(0,(start/1000.)),(SpecRange/4,(start/1000.+FreqRange/4000.)),(SpecRange/2,(start/1000.+FreqRange/2000.)),(3*SpecRange/4,(start/1000.+3*FreqRange/4000.)),(SpecRange,(start/1000.+FreqRange/1000.))]])

    def denoise_undo(self):
        # Listener for undo button in denoising dialog
        # TODO: Can I actually delete something from an object?
        print("Undoing",np.shape(self.audiodata_backup))
        if hasattr(self,'audiodata_backup'):
            if self.audiodata_backup is not None:
                if np.shape(self.audiodata_backup)[1]>0:
                    self.audiodata = np.copy(self.audiodata_backup[:,-1])
                    self.audiodata_backup = self.audiodata_backup[:,:-1]
                    self.sp.setNewData(self.audiodata,self.sampleRate)
                    sgRaw = self.sp.spectrogram(self.audiodata,self.sampleRate,mean_normalise=True,onesided=True,multitaper=False)
                    maxsg = np.min(sgRaw)
                    self.sg = np.abs(np.where(sgRaw == 0, 0.0, 10.0 * np.log10(sgRaw / maxsg)))
                    self.overviewImage.setImage(self.sg)
                    self.specPlot.setImage(self.sg)
                    self.amplPlot.setData(
                        np.linspace(0.0, float(self.datalength) / self.sampleRate, num=self.datalength, endpoint=True),
                        self.audiodata)
                    if hasattr(self,'seg'):
                        self.seg.setNewData(self.audiodata,sgRaw,self.sampleRate,self.config['window_width'],self.config['incr'])

                    self.setColourLevels()

                    #self.drawfigMain()

    def denoise_save(self):
        # Listener for save button in denoising dialog
        # Save denoised data
        # Other players need them to be 16 bit, which is this magic number
        #self.audiodata *= 32768.0
        #self.audiodata = self.audiodata.astype('int16')
        #import soundfile as sf
        filename = self.filename[:-4] + '_d' + self.filename[-4:]
        # self.audiodata*= 32768.0
        # print self.audiodata.dtype
        # self.audiodata = self.audiodata.astype('int16')
        #wavfile.write(filename,self.sampleRate, self.audiodata)
        wavio.write(filename,self.audiodata.astype('int16'),self.sampleRate,scale='dtype-limits', sampwidth=2) #sampwidth=2 for 'int16' output
        # librosa.output.write_wav(filename, self.audiodata, self.sampleRate)
        # msg = QMessageBox()
        # msg.setIcon(QMessageBox.Information)
        # msg.setWindowIcon(QIcon('Avianz.ico'))
        # msg.setText("Saved!")
        # msg.setWindowTitle("Update")
        # msg.setStandardButtons(QMessageBox.Ok)
        # msg.exec_()
        self.statusLeft.setText("Saved")

    def segmentationDialog(self):
        # Create the segmentation dialog when the relevant button is pressed
        self.segmentDialog = Dialogs.Segmentation(np.max(self.audiodata))
        self.segmentDialog.show()
        self.segmentDialog.activateWindow()
        self.segmentDialog.undo.clicked.connect(self.segment_undo)
        self.segmentDialog.activate.clicked.connect(self.segment)
        #self.segmentDialog.save.clicked.connect(self.segments_save)

    def segment(self):
        # Listener for the segmentation dialog
        # TODO: Currently just gives them all the label "Don't Know"'
        # TODO: Add in the wavelet one
        # TODO: More testing of the algorithms, parameters, etc.
        seglen = len(self.segments)
        [alg, ampThr, medThr,HarmaThr1,HarmaThr2,PowerThr,minfreq,minperiods,Yinthr,window,FIRThr1,species] = self.segmentDialog.getValues()
        #[alg, ampThr, medThr,HarmaThr1,HarmaThr2,PowerThr,minfreq,minperiods,Yinthr,window,FIRThr1,depth,thrType,thr,wavelet,bandchoice,start,end,species] = self.segmentDialog.getValues()
        species = str(species)
        #if not hasattr(self,'seg'):
        #    self.seg = Segment.Segment(self.audiodata,sgRaw,self.sp,self.sampleRate,self.config['minSegment'],self.config['window_width'],self.config['incr'])
        self.statusLeft.setText("Segmenting...")
        if str(alg) == "Amplitude":
            newSegments = self.seg.segmentByAmplitude(float(str(ampThr)))
        elif str(alg) == "Median Clipping":
            newSegments = self.seg.medianClip(float(str(medThr)))
            #print newSegments
        elif str(alg) == "Harma":
            newSegments = self.seg.Harma(float(str(HarmaThr1)),float(str(HarmaThr2)))
        elif str(alg) == "Power":
            newSegments = self.seg.segmentByPower(float(str(PowerThr)))
        elif str(alg) == "Onsets":
            newSegments = self.seg.onsets()
            #print newSegments
        elif str(alg) == "Fundamental Frequency":
            newSegments, pitch, times = self.seg.yin(int(str(minfreq)),int(str(minperiods)),float(str(Yinthr)),int(str(window)),returnSegs=True)
        elif str(alg) == "FIR":
            newSegments = self.seg.segmentByFIR(float(str(FIRThr1)))
            # print newSegments
        elif str(alg)=="Wavelets":
            newSegments = WaveletSegment.findCalls_test(fName=None,data=self.audiodata, sampleRate=self.sampleRate, species=species,trainTest=False)
        elif str(alg)=="Best":
            newSegments = self.seg.bestSegments()
            #print newSegments

            # # Here the idea is to use both ML and wavelets then label AND as definite and XOR as possible just for wavelets
            # # but ML is extremely slow and crappy. So I decided to use just the wavelets
            # newSegmentsML = WaveletSegment.findCalls_learn(fName=None,data=self.audiodata, sampleRate=self.sampleRate, species=species,trainTest=False)
            # print np.shape(newSegmentsML),type(newSegmentsML), newSegmentsML
            #
            # newSegments = WaveletSegment.findCalls_test(fName=None,data=self.audiodata, sampleRate=self.sampleRate, species='kiwi',trainTest=False)
            # # print type(newSegments),newSegments
            # import itertools
            # newSegments=list(itertools.chain.from_iterable(newSegments))
            # temp=np.zeros(len(newSegmentsML))
            # for i in newSegments:
            #     temp[i]=1
            # newSegments=temp.astype(int)
            # newSegments=newSegments.tolist()
            # print np.shape(newSegments), type(newSegments), newSegments
            #
            # newSegmentsDef=np.minimum.reduce([newSegmentsML,newSegments])
            # newSegmentsDef=newSegmentsDef.tolist()
            # print "newSegmentsDef:", np.shape(newSegmentsDef), type(newSegmentsDef), newSegmentsDef
            # C=[(a and not b) or (not a and b) for a,b in zip(newSegmentsML,newSegments)]
            # newSegmentsPb=[int(c) for c in C]
            # print "newSegmentsPosi:", np.shape(newSegmentsPb), type(newSegmentsPb), newSegmentsPb
            #
            # # convert these segments to [start,end] format
            # newSegmentsDef=self.binary2seg(newSegmentsDef)
            # newSegmentsPb=self.binary2seg(newSegmentsPb)

        # Generate annotation friendly output. That's enough for interface?
        # Merge neighbours for wavelet seg
        if str(alg)=="Wavelets":
            newSegments=self.mergeSeg(newSegments)
            for seg in newSegments:
                self.addSegment(float(seg[0]), float(seg[1]), 0, 0,
                                species.title() + "?")  # TODO: sometimes got index exceed max
        else:
            for seg in newSegments:
                self.addSegment(float(seg[0]),float(seg[1])) # TODO: sometimes got index exceed max

        # Save the excel file
        #print newSegments
        self.saveDetections(newSegments,mode='Excel',species=species)
        #     newSegmentsDef=self.mergeSeg(newSegmentsDef)
        #     newSegmentsPb=self.mergeSeg(newSegmentsPb)
        # for seg in newSegmentsDef:
        #     self.addSegment(float(seg[0]),float(seg[1]),0,0,species)
        # for seg in newSegmentsPb:
        #     self.addSegment(float(seg[0]),float(seg[1]),0,0,species+"?")

        # annotation=[]
        # for seg in newSegments:
        #     annotation.append([seg[0],seg[1],0,0,str(species)+"?"])
        # self.saveSegments2(annotation)

        self.lenNewSegments = len(newSegments)
        # # Generate binary output
        # # print "Binary output:"
        # n=math.ceil(float(self.datalength)/self.sampleRate)
        # # print 'n=', n
        # detected=np.zeros(int(n))
        # for seg in newSegments:
        #     for a in range(len(detected)):
        #         if math.floor(seg[0])<=a and a<math.ceil(seg[1]):
        #             detected[a]=1
        # # print detected
        # self.lenNewSegments = len(newSegments)
        # #Generate time stampls [start(mm:ss) end(mm:ss)]
        # # print "start end (mm:ss):"
        # annotation=[]
        # for seg in newSegments:
        #     annotation.append([self.convertMillisecs(seg[0]*1000),self.convertMillisecs(seg[1]*1000)])
        # # print annotation

        self.segmentDialog.undo.setEnabled(True)

        self.statusLeft.setText("Ready")

    def saveDetections(self, annotation, mode,species): # Origin: saveSegments from interface_FindSpecies
        # This saves the detections into three different formats: annotation, excel, and binary

        # method=self.algs.currentText()
        relfname = os.path.relpath(str(self.filename), str(self.dirName))
        eFile = self.dirName + '\DetectionSummary_' + species + '.xlsx'

        if mode == 'Annotation':
            if isinstance(self.filename, str):
                file = open(self.filename + '.data', 'w')
            else:
                file = open(str(self.filename) + '.data', 'w')
            json.dump(annotation, file)

        elif mode == 'Excel':
            if os.path.isfile(eFile):  # if the file is already there
                try:
                    wb = load_workbook(str(eFile))
                    ws = wb.get_sheet_by_name('TimeStamps')
                    c = 1
                    r = ws.max_row + 1  # TODO: get last row number from existing file
                    ws.cell(row=r, column=1, value=str(relfname))
                    for seg in annotation:
                        ws.cell(row=r, column=c + 1, value=str(seg[0]) + '-' + str(seg[1]))
                        c = c + 1
                    wb.save(str(eFile))
                except:
                    print "Unable to open file"  # Does not exist OR no read permissions
            else:
                wb = Workbook()
                wb.create_sheet(title='TimeStamps', index=1)
                wb.create_sheet(title='PresenceAbsence', index=2)
                wb.create_sheet(title='PerSecond', index=3)

                ws = wb.get_sheet_by_name('TimeStamps')
                ws.cell(row=1, column=1, value="File Name")
                ws.cell(row=1, column=2, value="Detections [start-end(mm:ss)]")
                c = 1
                r = 2
                ws.cell(row=r, column=c, value=str(relfname))
                for seg in annotation:
                    ws.cell(row=r, column=c + 1, value=str(seg[0]) + '-' + str(seg[1]))
                    c = c + 1
                # Second sheet
                ws = wb.get_sheet_by_name('PresenceAbsence')
                ws.cell(row=1, column=1, value="File Name")
                ws.cell(row=1, column=2, value="Presence/Absence")

                # Third sheet
                ws = wb.get_sheet_by_name('PerSecond')
                ws.cell(row=1, column=1, value="File Name")
                ws.cell(row=1, column=2, value="Presence=1/Absence=0")
                c = 2
                for i in range(900):
                    ws.cell(row=2, column=c, value="S " + str(i + 1))
                    c = c + 1
                first = wb.get_sheet_by_name('Sheet')
                wb.remove_sheet(first)
                wb.save(str(eFile))

            # Presence absence excel
            if os.path.isfile(eFile):  # if the file is already there
                try:
                    wb = load_workbook(str(eFile))
                    # ws=wb.create_sheet(title="PresenceAbsence",index=2)
                    ws = wb.get_sheet_by_name('PresenceAbsence')
                    r = ws.max_row + 1  #
                    ws.cell(row=r, column=1, value=str(relfname))
                    if annotation:
                        ws.cell(row=r, column=2, value='Yes')
                    else:
                        ws.cell(row=r, column=2, value='_')
                    wb.save(str(eFile))
                except:
                    print "Unable to open file"  # Does not exist OR no read permissions

        else:  # mode=='Binary'
            # eFile=self.dirName+'\\3PerSecond_'+self.species+'_'+'.xlsx'
            if os.path.isfile(eFile):  # if the file is already there
                try:
                    wb = load_workbook(str(eFile))
                    ws = wb.get_sheet_by_name('PerSecond')
                    c = 1
                    r = ws.max_row + 1  # TODO: get last row number from existing file
                    ws.cell(row=r, column=1, value=str(relfname))
                    for seg in annotation:
                        ws.cell(row=r, column=c + 1, value=seg)
                        c = c + 1
                    wb.save(str(eFile))
                except:
                    print "Unable to open file"  # Does not exist OR no read permissions

    def segment_undo(self):
        # Listener for undo button in segmentation dialog
        # This is very cheap: the segments were appended, so delete the last len of them (from the end)
        end = len(self.segments)
        for seg in range(end-1,end-self.lenNewSegments-1,-1):
            self.deleteSegment(seg)
        self.segmentDialog.undo.setEnabled(False)


    def findMatches(self):
        # Calls the cross-correlation function to find matches like the currently highlighted box
        # TODO: Other methods apart from c-c?
        # So there needs to be a currently highlighted box
        # TODO: If there isn't a box highlighted, grey out the menu option
        #if not hasattr(self,'seg'):
        #    self.seg = Segment.Segment(self.audiodata,sgRaw,self.sp,self.sampleRate,self.config['minSegment'],self.config['window_width'],self.config['incr'])

        if self.box1id is None or self.box1id == -1:
            print "No box selected"
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("No segment selected to match")
            msg.setWindowIcon(QIcon('img/Avianz.ico'))
            msg.setWindowTitle("No segment")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            return
        else:
            self.statusLeft.setText("Finding matches...")
            #[alg, thr] = self.matchDialog.getValues()
            # Only want to draw new segments, so find out how many there are now
            seglen = len(self.segments)
            # Get the segment -- note that takes the full y range
            if type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                x1 = self.listRectanglesa2[self.box1id].pos().x()
                x2 = x1 + self.listRectanglesa2[self.box1id].size().x()
            else:
                x1, x2 = self.listRectanglesa2[self.box1id].getRegion()
            #print x1, x2
            # Get the data for the spectrogram
            sgRaw = self.sp.spectrogram(self.audiodata,self.sampleRate,mean_normalise=True,onesided=True,multitaper=False)#***
            segment = sgRaw[int(x1):int(x2),:]
            len_seg = (x2-x1) * self.config['incr'] / self.sampleRate
            indices = self.seg.findCCMatches(segment,sgRaw,self.config['corrThr'])
            # indices are in spectrogram pixels, need to turn into times
            for i in indices:
                # Miss out the one selected: note the hack parameter
                if np.abs(i-x1) > self.config['overlap_allowed']:
                    time = float(i)*self.config['incr'] / self.sampleRate
                    self.addSegment(time, time+len_seg,0,0,self.segments[self.box1id][4])
            self.statusLeft.setText("Ready")

    def classifySegments(self):
        # Note that this still works on 1 second -- species-specific parameter eventually (here twice: as 1 and in sec loop)
        if self.segments is None or len(self.segments) == 0:
            print "No segments to recognise"
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setText("No segments to recognise")
            msg.setWindowIcon(QIcon('img/Avianz.ico'))
            msg.setWindowTitle("No segments")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            return
        else:
            # TODO: Ask for species; brown kiwi for now
            # TODO: ***** TIDY UP WAVELET SEG, USE THIS!
            for i in range(len(self.segments)):
                seglength = np.abs(self.segments[i][1] - self.segments[i][0])
                if seglength <= 1:
                    # Recognise as is
                    label = WaveletSegment.computeWaveletEnergy_1s(self.audiodata[self.segments[i][0]:self.segments[i][1]])
                    self.updateText(label,i)
                else:
                    for sec in range(np.ceil(seglength)):
                        label = WaveletSegment.computeWaveletEnergy_1s(self.audiodata[sec*self.sampleRate+self.segments[i][0]:(sec+1)*self.sampleRate+self.segments[i][0]])
                        # TODO: Check if the labels match, decide what to do if not
                    self.updateText(label,i)

    def recognise(self):
        # This will eventually call methods to do automatic recognition
        # Actually, will produce a dialog to ask which species, etc.
        # TODO
        pass

# ===============
# Code for playing sounds
    # These functions are the phonon playing code
    # Note that if want to play e.g. denoised one, will have to save it and then reload it
    def playSegment(self):
        self.segmentStop = self.playSlider.maximum()
        if self.media_obj.state() == phonon.Phonon.PlayingState:
            self.media_obj.pause()
            self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))
        elif self.media_obj.state() == phonon.Phonon.PausedState or self.media_obj.state() == phonon.Phonon.StoppedState:
            self.media_obj.play()
            self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPause))

    def playFinished(self):
        self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))
        self.media_obj.stop()
        self.media_obj.seek(self.playSlider.value()-self.widthWindow.value()*1000)
        # self.playSlider.setValue(self.convertAmpltoSpec(self.playSlider.value()/1000.0-self.widthWindow.value()))
        self.bar.setValue(self.convertAmpltoSpec(self.playSlider.value()/1000.0-self.widthWindow.value()))

    def sliderMoved(self):
        # When the slider is moved, change the position of playback
        self.media_obj.seek(self.playSlider.value())
        # playSlider.value() is in ms, need to convert this into spectrogram pixels
        self.bar.setValue(self.convertAmpltoSpec(self.playSlider.value()/1000.0))

    def barMoved(self,evt):
        self.playSlider.setValue(self.convertSpectoAmpl(evt.x())*1000)
        self.media_obj.seek(self.convertSpectoAmpl(evt.x())*1000)

    def movePlaySlider(self, time):
        if not self.playSlider.isSliderDown():
            self.playSlider.setValue(time)
        self.timePlayed.setText(self.convertMillisecs(time)+"/"+self.totalTime)
        if time > min(self.playSlider.maximum(),self.segmentStop):
            self.media_obj.stop()
            self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))
            self.media_obj.seek(self.playSlider.minimum())
        self.bar.setValue(self.convertAmpltoSpec(self.playSlider.value()/1000.0))

    def setSliderLimits(self, start,end):
        self.playSlider.setRange(start, end)
        self.playSlider.setValue(start)
        self.segmentStop = self.playSlider.maximum()
        self.media_obj.seek(start)

    def playSelectedSegment(self):
        # listener for PlaySegment button
        # Get selected segment start and end (or return if no segment selected)
        # This isn't pausable, since it goes back to the beginning. I think it's OK though -- they should be short?
        if self.box1id > -1:
            start = self.listRectanglesa1[self.box1id].getRegion()[0]*1000
            self.segmentStop = self.listRectanglesa1[self.box1id].getRegion()[1]*1000
            self.media_obj.seek(start)
            #self.media_obj.seek(start)
            #self.media_obj.play()
            #self.segmentStop = self.playSlider.maximum()
            if self.media_obj.state() == phonon.Phonon.PlayingState:
                self.media_obj.pause()
                # self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))
                # self.playSegButton.setIcon(QtGui.QIcon('img/playsegment.png'))
                # self.playButton.setText("Play")
            elif self.media_obj.state() == phonon.Phonon.PausedState or self.media_obj.state() == phonon.Phonon.StoppedState:
                self.media_obj.play()
                # self.playSegButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPause))

    def movePlaySlider2(self,time):
        if not self.playSlider.isSliderDown():
            self.playSlider.setValue(time)
        #self.timePlayed.setText(self.convertMillisecs(time)+"/"+self.totalTime)
        if time > min(self.playSlider.maximum(),self.segmentStop):
            self.media_obj2.stop()
            #self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))
            #self.media_obj.seek(self.playSlider.minimum())
        self.bar2.setValue(self.convertAmpltoSpec(self.playSlider.value()/1000.0)+self.bandLimitedStart)

    def playFinished2(self):
        self.media_obj2.stop()
        self.p_spec.removeItem(self.bar2)

    def playBandLimitedSegment(self):
        # listener for PlayBandlimitedSegment button
        # Get the band limits of the segment, bandpass filter, then play that
        start = int(self.listRectanglesa1[self.box1id].getRegion()[0]*self.sampleRate)
        stop = int(self.listRectanglesa1[self.box1id].getRegion()[1]*self.sampleRate)
        bottom = int(self.segments[self.box1id][2]*self.sampleRate/2./np.shape(self.sg)[1])
        top = int(self.segments[self.box1id][3]*self.sampleRate/2./np.shape(self.sg)[1])

        if bottom > 0 and top>0:
            self.bandLimitedStart=self.convertAmpltoSpec(float(start)/self.sampleRate)
            data = self.audiodata[start:stop]
            data=self.sp.bandpassFilter(data,bottom,top)
            # data = self.sp.ButterworthBandpass(data, self.sampleRate, bottom, top,order=5)

            if platform.system() == 'Darwin':
                filename = 'temp.wav'
            else:
                import tempfile
                f = tempfile.NamedTemporaryFile(mode='w+t', delete=False)
                filename = f.name
            data = data.astype('int16')
            wavio.write(filename,data,self.sampleRate,scale='dtype-limits',sampwidth=2)

            if not hasattr(self, 'bar2'):
                self.bar2 = pg.InfiniteLine(angle=90, movable=True, pen={'color': 'r', 'width': 2})
                self.media_obj2 = phonon.Phonon.MediaObject(self)
                audio_output = phonon.Phonon.AudioOutput(phonon.Phonon.MusicCategory, self)
                phonon.Phonon.createPath(self.media_obj2, audio_output)
                self.media_obj2.setTickInterval(20)
                self.media_obj2.tick.connect(self.movePlaySlider2)
                self.media_obj2.finished.connect(self.playFinished2)
            self.bar2.setValue(self.bandLimitedStart)
            self.p_spec.addItem(self.bar2, ignoreBounds=True)

            # Instantiate a Qt media object and prepare it (for audio playback)
            self.media_obj2.setCurrentSource(phonon.Phonon.MediaSource(filename))
            self.media_obj2.seek(0)
            self.media_obj2.play()
            #f.close()
            return
        else:
            self.playSelectedSegment()

# ============
# Various actions: deleting, saving, quitting
    def deleteSegment(self,id=-1):
        """ Listener for delete segment button, or backspace key. Also called when segments are deleted by the
        human classify dialogs.
        Deletes the segment that is selected, otherwise does nothing.
        Updates the overview segments as well.
        """
        if id<0:
            #
            id = self.box1id
        if id>-1:
            # Work out which overview segment this segment is in (could be more than one) and update it
            inds = int(float(self.convertAmpltoSpec(self.segments[id][0]))/self.widthOverviewSegment)
            inde = int(float(self.convertAmpltoSpec(self.segments[id][1]))/self.widthOverviewSegment)

            if self.segments[id][4] == "Don't Know":
                self.overviewSegments[inds:inde+1,0] -= 1
            elif self.segments[id][4][-1] == '?':
                self.overviewSegments[inds:inde + 1, 2] -= 1
            else:
                self.overviewSegments[inds:inde + 1, 1] -= 1
            for box in range(inds, inde + 1):
                if self.overviewSegments[box,0] > 0:
                    self.SegmentRects[box].setBrush(self.ColourNone)
                elif self.overviewSegments[box,2] > 0:
                    self.SegmentRects[box].setBrush(self.ColourPossible)
                elif self.overviewSegments[box,1] > 0:
                    self.SegmentRects[box].setBrush(self.ColourNamed)
                else:
                    self.SegmentRects[box].setBrush(pg.mkBrush('w'))

            self.p_ampl.removeItem(self.listRectanglesa1[id])
            self.p_spec.removeItem(self.listRectanglesa2[id])
            self.p_spec.removeItem(self.listLabels[id])
            del self.listLabels[id]
            del self.segments[id]
            del self.listRectanglesa1[id]
            del self.listRectanglesa2[id]
            self.box1id = -1

    # def chDir(self):
    #     # Listener for Change directory menu item
    #     # Don't want to change what's drawn there (currently selected file and its spectro),
    #     # but effects when opening a new file
    #     dir= QtGui.QFileDialog.getExistingDirectory(self,'Choose Directory',self.dirName,QtGui.QFileDialog.ShowDirsOnly) #"Wav files (*.wav)")
    #     dir = str(dir)
    #     if dir!='':
    #         self.dirName=os.path.abspath(dir)
    #         # Now repopulate the listbox
    #         #print "Now in "+self.listOfFiles[i].fileName()
    #         # self.dirName=str(dir.absolutePath())
    #         self.listFiles.clearSelection()
    #         self.listFiles.clearFocus()
    #         self.listFiles.clear()
    #         self.previousFile = None
    #         self.fillFileList(fileName=None)

    def setOperatorReviewerDialog(self):
        # Create the Operator/reviewer dialog when the menu item is selected
    #     Listener for Set Operator/Reviewer menu item
    #     if not hasattr(self,'setOperatorReviewerDialog'):
        self.setOperatorReviewerDialog = Dialogs.OperatorReviewer()
        self.setOperatorReviewerDialog.show()
        self.setOperatorReviewerDialog.activateWindow()
        self.setOperatorReviewerDialog.activate.clicked.connect(self.operator)

    def operator(self):
        # Listener for the operator/reviewer dialog.
        # TODO: save these
        [isOperator,name] = self.setOperatorReviewerDialog.getValues()
        if isOperator==True:
            self.operator=name
            self.statusRight.setText("Operator: "+self.operator)
        else:
            self.reviewer=name
            self.statusRight.setText("Reviewer: "+self.reviewer)

    def deleteAll(self):
        """ Listener for delete all button.
        Checks if the user meant to do it, then calls removeSegments()
        """
        reply = QMessageBox.question(self,"Delete All Segments","Are you sure you want to delete all segments?",    QMessageBox.Yes | QMessageBox.No)
        # reply.setWindowIcon(QIcon('Avianz.ico'))
        if reply==QMessageBox.Yes:
            self.removeSegments()

    def removeSegments(self):
        """ Remove all the segments in response to the menu selection, or when a new file is loaded. """
        for r in self.listLabels:
            self.p_spec.removeItem(r)
        for r in self.listRectanglesa1:
            self.p_ampl.removeItem(r)
        for r in self.listRectanglesa2:
            self.p_spec.removeItem(r)
        for r in self.SegmentRects:
            r.setBrush(pg.mkBrush('w'))
            r.update()

        self.segments=[]
        self.listRectanglesa1 = []
        self.listRectanglesa2 = []
        self.listLabels = []
        self.box1id = -1

    def saveSegments(self):
        """ Save the segmentation data as a json file.
        Name of the file is the name of the wave file + .data"""
        if len(self.segments)>0 or self.hasSegments:
            print("Saving segments to "+self.filename)
            if isinstance(self.filename, str):
                file = open(self.filename + '.data', 'w')
            else:
                file = open(str(self.filename) + '.data', 'w')
            json.dump(self.segments,file)

    def mergeSeg(self,segments):
        indx=[]
        for i in range(len(segments)-1):
            if segments[i][1]==segments[i+1][0]:
                indx.append(i)
        indx.reverse()
        for i in indx:
            segments[i][1]=segments[i+1][1]
            del(segments[i+1])
        return segments

    def binary2seg(self,binary):
        segments=[]
        for i in range(len(binary)):
            if binary[i]==1:
                segments.append([i,i+1])
        return segments

    def closeEvent(self, event):
        """ Catch the user closing the window by clicking the Close button instead of quitting. """
        self.quit()

    def quit(self):
        """ Listener for the quit button, also called by closeEvent().
        Add in the username at the top, and then save the segments and the config file.
        """
        print("Quitting")
        if len(self.segments) > 0:
            if self.segments[0][0] > -1:
                self.segments.insert(0, [-1, -1, str(self.username), -1, -1])
        else:
            self.segments.insert(0, [-1, -1, str(self.username), -1, -1])
        self.saveSegments()
        if self.saveConfig == True:
            print "Saving config file"
            json.dump(self.config, open(self.configfile, 'wb'))
        QApplication.quit()

# =============
# Helper functions

    def splitFile5mins(self, name):
        # Nirosha wants to split files that are long (15 mins) into 5 min segments
        # Could be used when loading long files :)
        try:
            self.audiodata, self.sampleRate = lr.load(name,sr=None)
        except:
            print("Error: try another file")
        nsamples = np.shape(self.audiodata)[0]
        lengthwanted = self.sampleRate * 60 * 5
        count = 0
        while (count + 1) * lengthwanted < nsamples:
            data = self.audiodata[count * lengthwanted:(count + 1) * lengthwanted]
            filename = name[:-4] + '_' +str(count) + name[-4:]
            lr.output.write_wav(filename, data, self.sampleRate)
            count += 1
        data = self.audiodata[(count) * lengthwanted:]
        filename = name[:-4] + '_' + str((count)) + name[-4:]
        lr.output.write_wav(filename,data,self.sampleRate)

# Start the application
app = QApplication(sys.argv)

DOC=True    # DOC features or all

# This screen asks what you want to do, then processes the response
first = Dialogs.StartScreen(DOC=DOC)
first.setWindowIcon(QtGui.QIcon('img/AviaNZ.ico'))
first.show()
app.exec_()

task = first.getValues()

if task == 1:
    if DOC and False:
        # Modal dialog to get the user data
        userdata = Dialogs.getUserData()
        userdata.show()
        userdata.exec_()
        username = userdata.getValues()
        avianz = AviaNZ(configfile='AviaNZconfig.txt',username=username,DOC=DOC)
    else:
        avianz = AviaNZ(configfile='AviaNZconfig.txt',username=None,DOC=DOC)
    avianz.setWindowIcon(QtGui.QIcon('img/AviaNZ.ico'))
    avianz.show()
    app.exec_()
elif task==2:
    avianz = interface_FindSpecies.AviaNZFindSpeciesInterface()
    avianz.setWindowIcon(QtGui.QIcon('img/AviaNZ.ico'))
    avianz.show()
    app.exec_()
